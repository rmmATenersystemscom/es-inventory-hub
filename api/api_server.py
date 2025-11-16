#!/usr/bin/env python3
"""
ES Inventory Hub API Server

Provides REST API endpoints for:
1. Accessing variance report data
2. Triggering collector runs
3. Checking system status

Usage:
    python3 api_server.py
    # Server runs on https://db-api.enersystems.com:5400
"""

import os
import sys
import json
import subprocess
import csv
import io
from datetime import date, datetime, timedelta
from typing import Dict, List, Any, Optional

from flask import Flask, jsonify, request, Response
from flask_cors import CORS
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker

# Export functionality imports
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Add the project root to Python path
sys.path.insert(0, '/opt/es-inventory-hub')

from collectors.checks.cross_vendor import run_cross_vendor_checks

app = Flask(__name__)

# Import and register QBR API blueprint
from api.qbr_api import qbr_api
app.register_blueprint(qbr_api)

def _format_date_string(date_str):
    """Format date string to consistent ISO 8601 format without microseconds"""
    if not date_str or date_str == 'Unknown':
        return 'Unknown'
    
    try:
        # Parse the date string and reformat without microseconds
        if isinstance(date_str, str):
            # Handle various date formats
            if 'T' in date_str:
                # ISO format - remove microseconds and timezone offset
                dt = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                return dt.replace(microsecond=0).replace(tzinfo=None).isoformat() + 'Z'
            else:
                return date_str
        return date_str
    except (ValueError, TypeError):
        return date_str

# Enable CORS for cross-origin requests with specific dashboard access
CORS(app, 
     origins=['https://dashboards.enersystems.com', 'http://localhost:3000', 'http://localhost:8080'],
     methods=['GET', 'POST', 'PUT', 'DELETE', 'OPTIONS'],
     allow_headers=['Content-Type', 'Authorization', 'X-Requested-With', 'X-API-Key', 'Cache-Control', 'Pragma'],
     supports_credentials=True,
     max_age=86400)

# Database connection
from common.config import get_dsn
DSN = get_dsn()
engine = create_engine(DSN)
Session = sessionmaker(bind=engine)

def get_session():
    """Get database session."""
    return Session()

def get_latest_matching_date() -> Optional[date]:
    """Get the latest date where both vendors have data."""
    with get_session() as session:
        query = text("""
            SELECT snapshot_date, COUNT(DISTINCT vendor_id) as vendor_count
            FROM device_snapshot
            WHERE snapshot_date >= CURRENT_DATE - INTERVAL '7 days'
            GROUP BY snapshot_date
            HAVING COUNT(DISTINCT vendor_id) = 2
            ORDER BY snapshot_date DESC
            LIMIT 1
        """)
        
        result = session.execute(query).fetchone()
        return result[0] if result else None

def get_organization_breakdown(session, exceptions, report_date):
    """Get detailed breakdown of exceptions by organization for each variance type."""
    # Get organization data for each exception
    org_query = text("""
        SELECT 
            e.id,
            e.type,
            e.hostname,
            e.details,
            e.resolved,
            COALESCE(
                ds_tl.organization_name,
                ds_ninja.organization_name,
                'Unknown'
            ) as organization_name,
            COALESCE(
                ds_tl.display_name,
                ds_ninja.display_name,
                e.hostname
            ) as display_name,
            COALESCE(
                ds_tl.device_status,
                ds_ninja.device_status,
                'Unknown'
            ) as billing_status,
            CASE 
                WHEN v_tl.name = 'ThreatLocker' THEN 'ThreatLocker'
                WHEN v_ninja.name = 'Ninja' THEN 'Ninja'
                ELSE 'Unknown'
            END as vendor
        FROM exceptions e
        LEFT JOIN device_snapshot ds_tl ON ds_tl.hostname = e.hostname 
            AND ds_tl.snapshot_date = :report_date
            AND ds_tl.vendor_id = (SELECT id FROM vendor WHERE name = 'ThreatLocker')
        LEFT JOIN vendor v_tl ON ds_tl.vendor_id = v_tl.id
        LEFT JOIN device_snapshot ds_ninja ON ds_ninja.hostname = e.hostname 
            AND ds_ninja.snapshot_date = :report_date
            AND ds_ninja.vendor_id = (SELECT id FROM vendor WHERE name = 'Ninja')
        LEFT JOIN vendor v_ninja ON ds_ninja.vendor_id = v_ninja.id
        WHERE e.date_found = :report_date
        ORDER BY e.type, organization_name, e.hostname
    """)
    
    org_results = session.execute(org_query, {'report_date': report_date}).fetchall()
    
    # Group by type and organization
    by_organization = {}
    
    for row in org_results:
        exc_id, exc_type, hostname, details, resolved, org_name, display_name, billing_status, vendor = row
        
        # Map exception types to dashboard format
        dashboard_type = None
        if exc_type == "MISSING_NINJA":
            dashboard_type = "missing_in_ninja"
        elif exc_type == "DUPLICATE_TL":
            dashboard_type = "threatlocker_duplicates"
        elif exc_type == "SPARE_MISMATCH":
            dashboard_type = "devices_that_should_not_have_threatlocker"
        elif exc_type == "DISPLAY_NAME_MISMATCH":
            dashboard_type = "display_name_mismatches"
        
        if not dashboard_type:
            continue
            
        if dashboard_type not in by_organization:
            by_organization[dashboard_type] = {
                "total_count": 0,
                "by_organization": {}
            }
        
        # For Missing in Ninja, ThreatLocker Duplicates, and DevicesThatShouldNotHaveThreatlocker, extract organization from details if not found in device_snapshot
        if exc_type in ["MISSING_NINJA", "DUPLICATE_TL", "SPARE_MISMATCH"] and org_name == "Unknown" and details:
            # Extract organization from the details JSONB field
            details_dict = details if isinstance(details, dict) else {}
            
            # For SPARE_MISMATCH (DevicesThatShouldNotHaveThreatlocker), prefer ninja_org_name over tl_org_name
            if exc_type == "SPARE_MISMATCH":
                ninja_org = details_dict.get('ninja_org_name', '')
                tl_org = details_dict.get('tl_org_name', '')
                
                # Use ninja_org if it exists and is not empty, otherwise use tl_org, otherwise 'Unknown'
                if ninja_org and ninja_org.strip():
                    org_name = ninja_org
                    vendor = 'Ninja'
                elif tl_org and tl_org.strip():
                    org_name = tl_org
                    vendor = 'ThreatLocker'
                else:
                    org_name = 'Unknown'
            else:
                # For MISSING_NINJA and DUPLICATE_TL, use tl_org_name
                org_name = details_dict.get('tl_org_name', 'Unknown')
                if details_dict.get('tl_org_name'):
                    vendor = 'ThreatLocker'
            
            # Update display_name from site information
            if details_dict.get('tl_site_name'):
                display_name = f"{hostname} ({details_dict.get('tl_site_name')})"
            elif details_dict.get('ninja_site'):
                display_name = f"{hostname} ({details_dict.get('ninja_site')})"
        
        # For Display Name Mismatches, extract display name information from details
        if exc_type == "DISPLAY_NAME_MISMATCH" and details:
            details_dict = details if isinstance(details, dict) else {}
            
            # Extract organization from details (prefer ninja_org_name)
            ninja_org = details_dict.get('ninja_org_name', '')
            tl_org = details_dict.get('tl_org_name', '')
            
            if ninja_org and ninja_org.strip():
                org_name = ninja_org
                vendor = 'Ninja'
            elif tl_org and tl_org.strip():
                org_name = tl_org
                vendor = 'ThreatLocker'
            
            # Use the Ninja display name as the primary display name for the modal
            ninja_display = details_dict.get('ninja_display_name', '')
            if ninja_display and ninja_display.strip():
                display_name = ninja_display
        
        if org_name not in by_organization[dashboard_type]["by_organization"]:
            by_organization[dashboard_type]["by_organization"][org_name] = []
        
        # Create device entry
        device_entry = {
            "hostname": hostname,
            "vendor": vendor,
            "display_name": display_name,
            "organization": org_name,
            "billing_status": billing_status,
            "action": "Investigate"
        }
        
        # For Display Name Mismatches, add specific display name information for the modal
        if exc_type == "DISPLAY_NAME_MISMATCH" and details:
            details_dict = details if isinstance(details, dict) else {}
            device_entry.update({
                "ninja_display_name": details_dict.get('ninja_display_name', ''),
                "threatlocker_display_name": details_dict.get('tl_display_name', ''),
                "ninja_hostname": details_dict.get('ninja_hostname', ''),
                "threatlocker_hostname": details_dict.get('tl_hostname', '')
            })
        
        by_organization[dashboard_type]["by_organization"][org_name].append(device_entry)
        by_organization[dashboard_type]["total_count"] += 1
    
    return by_organization

def get_data_status() -> Dict[str, Any]:
    """Determine the status of the data (current, stale, out_of_sync)."""
    latest_date = get_latest_matching_date()
    
    if not latest_date:
        return {
            "status": "out_of_sync",
            "message": "No matching data found between vendors",
            "latest_date": None
        }
    
    days_old = (date.today() - latest_date).days
    
    if days_old <= 1:
        return {
            "status": "current",
            "message": "Data is current",
            "latest_date": latest_date.isoformat()
        }
    else:
        return {
            "status": "stale_data",
            "message": f"Data is {days_old} days old",
            "latest_date": latest_date.isoformat()
        }

# API Endpoints

@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint."""
    return jsonify({
        "status": "healthy",
        "timestamp": datetime.now().isoformat(),
        "version": "1.0.0"
    })

@app.route('/api/status', methods=['GET'])
def get_status():
    """Get overall system status with collector health and data freshness information."""
    data_status = get_data_status()
    
    with get_session() as session:
        # Auto-cleanup stale jobs when status is checked
        cleanup_stale_jobs(session)
        
        # Get device counts and latest snapshot dates per vendor
        # Only count devices from each vendor's latest snapshot date
        device_query = text("""
            SELECT 
                v.name as vendor,
                COUNT(ds.id) as count,
                MAX(ds.snapshot_date) as latest_date
            FROM vendor v
            LEFT JOIN device_snapshot ds ON v.id = ds.vendor_id
                AND ds.snapshot_date = (
                    SELECT MAX(snapshot_date)
                    FROM device_snapshot ds2
                    WHERE ds2.vendor_id = v.id
                )
            WHERE v.name IN ('Ninja', 'ThreatLocker')
            GROUP BY v.name
            ORDER BY v.name
        """)
        
        device_results = session.execute(device_query).fetchall()
        device_counts = {}
        vendor_status = {}
        
        for row in device_results:
            vendor_name = row[0]
            count = int(row[1]) if row[1] else 0
            latest_date = row[2]
            
            device_counts[vendor_name] = count
            
            # Determine vendor data freshness
            if latest_date:
                days_old = (date.today() - latest_date).days
                if days_old == 0:
                    freshness_status = "current"
                    freshness_message = "Data is current"
                elif days_old == 1:
                    freshness_status = "yesterday"
                    freshness_message = "Data is from yesterday"
                elif days_old <= 3:
                    freshness_status = "stale"
                    freshness_message = f"Data is {days_old} days old"
                else:
                    freshness_status = "very_stale"
                    freshness_message = f"Data is {days_old} days old"
            else:
                latest_date = None
                freshness_status = "no_data"
                freshness_message = "No data available"
            
            vendor_status[vendor_name] = {
                "latest_date": latest_date.isoformat() if latest_date else None,
                "freshness_status": freshness_status,
                "freshness_message": freshness_message,
                "days_old": days_old if latest_date else None
            }
        
        # Ensure both vendors are always present (default to 0 if missing)
        if 'Ninja' not in device_counts:
            device_counts['Ninja'] = 0
            vendor_status['Ninja'] = {
                "latest_date": None,
                "freshness_status": "no_data",
                "freshness_message": "No data available",
                "days_old": None
            }
        if 'ThreatLocker' not in device_counts:
            device_counts['ThreatLocker'] = 0
            vendor_status['ThreatLocker'] = {
                "latest_date": None,
                "freshness_status": "no_data",
                "freshness_message": "No data available",
                "days_old": None
            }
        
        # Get recent collector failures (last 24 hours)
        collector_health_query = text("""
            SELECT 
                job_name,
                status,
                message,
                error,
                started_at,
                ended_at
            FROM job_runs
            WHERE job_name IN ('ninja-collector', 'threatlocker-collector')
            AND started_at >= NOW() - INTERVAL '24 hours'
            AND status = 'failed'
            ORDER BY started_at DESC
            LIMIT 5
        """)
        
        recent_failures = []
        for row in session.execute(collector_health_query).fetchall():
            job_name, status, message, error, started_at, ended_at = row
            collector_name = job_name.replace('-collector', '').title()
            recent_failures.append({
                "collector": collector_name,
                "job_name": job_name,
                "status": status,
                "message": message,
                "error": error,
                "started_at": started_at.isoformat() + 'Z' if started_at else None,
                "ended_at": ended_at.isoformat() + 'Z' if ended_at else None
            })
        
        # Check if any vendor has stale data or failed collectors
        has_warnings = False
        warnings = []
        
        for vendor_name, status_info in vendor_status.items():
            if status_info['freshness_status'] in ['stale', 'very_stale', 'no_data']:
                has_warnings = True
                warnings.append(f"{vendor_name} data is {status_info['freshness_message'].lower()}")
        
        if recent_failures:
            has_warnings = True
            failed_collectors = [f['collector'] for f in recent_failures]
            warnings.append(f"Recent collector failures: {', '.join(set(failed_collectors))}")
        
        # Get exception counts
        exception_query = text("""
            SELECT type, COUNT(*) as count
            FROM exceptions
            WHERE resolved = FALSE
            AND date_found = CURRENT_DATE
            GROUP BY type
        """)
        
        exception_counts = {row[0]: row[1] for row in session.execute(exception_query).fetchall()}
    
    return jsonify({
        "data_status": data_status,
        "device_counts": device_counts,
        "vendor_status": vendor_status,
        "collector_health": {
            "recent_failures": recent_failures,
            "has_recent_failures": len(recent_failures) > 0,
            "total_failures_last_24h": len(recent_failures)
        },
        "warnings": warnings if has_warnings else [],
        "has_warnings": has_warnings,
        "exception_counts": exception_counts,
        "total_exceptions": sum(exception_counts.values())
    })

@app.route('/api/variance-report/latest', methods=['GET'])
def get_latest_variance_report():
    """Get the latest variance report."""
    # Get query parameters
    include_resolved = request.args.get('include_resolved', 'false').lower() == 'true'
    
    latest_date = get_latest_matching_date()
    
    if not latest_date:
        return jsonify({
            "error": "No matching data found between vendors",
            "status": "out_of_sync"
        }), 400
    
    with get_session() as session:
        # Get exceptions for the latest date (filter resolved by default)
        if include_resolved:
            query = text("""
                SELECT id, date_found, type, hostname, details, resolved
                FROM exceptions
                WHERE date_found = :report_date
                ORDER BY type, hostname
            """)
        else:
            query = text("""
                SELECT id, date_found, type, hostname, details, resolved
                FROM exceptions
                WHERE date_found = :report_date
                AND resolved = FALSE
                ORDER BY type, hostname
            """)
        
        exceptions = session.execute(query, {'report_date': latest_date}).fetchall()
        
        # Group by type
        by_type = {}
        for exc in exceptions:
            exc_type = exc[2]
            if exc_type not in by_type:
                by_type[exc_type] = []
            by_type[exc_type].append({
                'id': exc[0],
                'hostname': exc[3],
                'details': exc[4],
                'resolved': exc[5]
            })
        
        # Calculate totals
        total_exceptions = len(exceptions)
        unresolved_count = sum(1 for exc in exceptions if not exc[5])
        
        # Create dashboard-compatible format
        exception_counts = {exc_type: len(devices) for exc_type, devices in by_type.items()}
        
        # Map to dashboard expected format
        dashboard_format = {
            "missing_in_ninja": {
                "total_count": exception_counts.get("MISSING_NINJA", 0)
            },
            "threatlocker_duplicates": {
                "total_count": exception_counts.get("DUPLICATE_TL", 0)
            },
            "devices_that_should_not_have_threatlocker": {
                "total_count": exception_counts.get("SPARE_MISMATCH", 0)
            },
            "display_name_mismatches": {
                "total_count": exception_counts.get("DISPLAY_NAME_MISMATCH", 0)
            }
        }
        
        # Get collection timestamps
        collection_info = _get_collection_timestamps(session, latest_date)
        
        # Get detailed organization breakdown
        org_breakdown = get_organization_breakdown(session, exceptions, latest_date)
        
        # Update dashboard format with organization data
        enhanced_dashboard_format = {
            "missing_in_ninja": org_breakdown.get("missing_in_ninja", {
                "total_count": exception_counts.get("MISSING_NINJA", 0),
                "by_organization": {}
            }),
            "threatlocker_duplicates": org_breakdown.get("threatlocker_duplicates", {
                "total_count": exception_counts.get("DUPLICATE_TL", 0),
                "by_organization": {}
            }),
            "devices_that_should_not_have_threatlocker": org_breakdown.get("devices_that_should_not_have_threatlocker", {
                "total_count": exception_counts.get("SPARE_MISMATCH", 0),
                "by_organization": {}
            }),
            "display_name_mismatches": org_breakdown.get("display_name_mismatches", {
                "total_count": exception_counts.get("DISPLAY_NAME_MISMATCH", 0),
                "by_organization": {}
            })
        }
        
        return jsonify({
            "report_date": latest_date.isoformat(),
            "data_status": get_data_status(),
            "summary": {
                "total_exceptions": total_exceptions,
                "unresolved_count": unresolved_count,
                "resolved_count": total_exceptions - unresolved_count
            },
            "exceptions_by_type": by_type,
            "exception_counts": exception_counts,
            # Collection information
            "collection_info": {
                "ninja_collected": collection_info["ninja_collected"],
                "threatlocker_collected": collection_info["threatlocker_collected"],
                "last_collection": collection_info["last_collection"],
                "data_freshness": collection_info["data_freshness"]
            },
            # Enhanced dashboard-compatible format with organization breakdown
            **enhanced_dashboard_format
        })

@app.route('/api/variance-report/<date_str>', methods=['GET'])
def get_variance_report_by_date(date_str: str):
    """Get variance report for a specific date."""
    # Get query parameters
    include_resolved = request.args.get('include_resolved', 'false').lower() == 'true'
    
    try:
        report_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({"error": "Invalid date format. Use YYYY-MM-DD"}), 400
    
    with get_session() as session:
        # Check if data exists for this date
        check_query = text("""
            SELECT COUNT(DISTINCT vendor_id) as vendor_count
            FROM device_snapshot
            WHERE snapshot_date = :report_date
        """)
        
        result = session.execute(check_query, {'report_date': report_date}).fetchone()
        
        if result[0] < 2:
            return jsonify({
                "error": f"Insufficient data for {date_str}. Found {result[0]} vendors, need 2.",
                "status": "insufficient_data"
            }), 400
        
        # Get exceptions for the date (filter resolved by default)
        if include_resolved:
            query = text("""
                SELECT id, date_found, type, hostname, details, resolved
                FROM exceptions
                WHERE date_found = :report_date
                ORDER BY type, hostname
            """)
        else:
            query = text("""
                SELECT id, date_found, type, hostname, details, resolved
                FROM exceptions
                WHERE date_found = :report_date
                AND resolved = FALSE
                ORDER BY type, hostname
            """)
        
        exceptions = session.execute(query, {'report_date': report_date}).fetchall()
        
        # Group by type
        by_type = {}
        for exc in exceptions:
            exc_type = exc[2]
            if exc_type not in by_type:
                by_type[exc_type] = []
            by_type[exc_type].append({
                'id': exc[0],
                'hostname': exc[3],
                'details': exc[4],
                'resolved': exc[5]
            })
        
        # Get collection timestamps
        collection_info = _get_collection_timestamps(session, report_date)
        
        return jsonify({
            "report_date": report_date.isoformat(),
            "summary": {
                "total_exceptions": len(exceptions),
                "unresolved_count": sum(1 for exc in exceptions if not exc[5]),
                "resolved_count": sum(1 for exc in exceptions if exc[5])
            },
            "exceptions_by_type": by_type,
            "exception_counts": {exc_type: len(devices) for exc_type, devices in by_type.items()},
            "collection_info": collection_info
        })

@app.route('/api/collectors/run', methods=['POST'])
def run_collectors():
    """Trigger collector runs with batch and job tracking."""
    import uuid
    from datetime import datetime, timedelta
    
    data = request.get_json() or {}
    collectors = data.get('collectors', ['ninja', 'threatlocker'])
    priority = data.get('priority', 'normal')
    run_cross_vendor = data.get('run_cross_vendor', True)
    
    # Generate unique IDs
    batch_id = f"bc_{uuid.uuid4().hex[:8]}"
    now = datetime.utcnow()
    
    try:
        with get_session() as session:
            # Create batch record
            batch_query = text("""
                INSERT INTO job_batches (batch_id, created_at, status, priority, started_at, message)
                VALUES (:batch_id, :created_at, :status, :priority, :started_at, :message)
            """)
            session.execute(batch_query, {
                'batch_id': batch_id,
                'created_at': now,
                'status': 'queued',
                'priority': priority,
                'started_at': now,
                'message': f"Starting collectors: {', '.join(collectors)}"
            })
            
            # Create job records - only for actual collectors (ninja, threatlocker)
            job_runs = []
            for collector in collectors:
                # Only create jobs for actual collectors, not analysis jobs
                if collector in ['ninja', 'threatlocker']:
                    job_id = f"{collector[:2]}_{uuid.uuid4().hex[:8]}"
                    job_name = f"{collector}-collector"
                    
                    job_query = text("""
                        INSERT INTO job_runs (job_id, batch_id, job_name, status, started_at, updated_at, message)
                        VALUES (:job_id, :batch_id, :job_name, :status, :started_at, :updated_at, :message)
                    """)
                    session.execute(job_query, {
                        'job_id': job_id,
                        'batch_id': batch_id,
                        'job_name': job_name,
                        'status': 'queued',
                        'started_at': now,
                        'updated_at': now,
                        'message': f"Queued {collector} collector"
                    })
                    
                    job_runs.append({
                        'job_name': job_name,
                        'job_id': job_id,
                        'status': 'queued',
                        'started_at': now.isoformat() + 'Z'
                    })
            
            # Add cross-vendor checks job if requested
            if run_cross_vendor:
                cross_vendor_job_id = f"cv_{uuid.uuid4().hex[:8]}"
                cross_vendor_job_name = "cross-vendor-checks"
                
                job_query = text("""
                    INSERT INTO job_runs (job_id, batch_id, job_name, status, started_at, updated_at, message)
                    VALUES (:job_id, :batch_id, :job_name, :status, :started_at, :updated_at, :message)
                """)
                session.execute(job_query, {
                    'job_id': cross_vendor_job_id,
                    'batch_id': batch_id,
                    'job_name': cross_vendor_job_name,
                    'status': 'queued',
                    'started_at': now,
                    'updated_at': now,
                    'message': "Queued cross-vendor consistency checks"
                })
                
                job_runs.append({
                    'job_name': cross_vendor_job_name,
                    'job_id': cross_vendor_job_id,
                    'status': 'queued',
                    'started_at': now.isoformat() + 'Z'
                })
            
            # Add Windows 11 24H2 assessment job
            windows_24h2_job_id = f"w24_{uuid.uuid4().hex[:8]}"
            windows_24h2_job_name = "windows-11-24h2-assessment"
            
            job_query = text("""
                INSERT INTO job_runs (job_id, batch_id, job_name, status, started_at, updated_at, message)
                VALUES (:job_id, :batch_id, :job_name, :status, :started_at, :updated_at, :message)
            """)
            session.execute(job_query, {
                'job_id': windows_24h2_job_id,
                'batch_id': batch_id,
                'job_name': windows_24h2_job_name,
                'status': 'queued',
                'started_at': now,
                'updated_at': now,
                'message': "Queued Windows 11 24H2 assessment"
            })
            
            job_runs.append({
                'job_name': windows_24h2_job_name,
                'job_id': windows_24h2_job_id,
                'status': 'queued',
                'started_at': now.isoformat() + 'Z'
            })
            
            session.commit()
            
            # Execute collectors INDEPENDENTLY - don't cancel one if another fails
            # Load environment variables from primary source for subprocess
            import subprocess as sp_module
            env = os.environ.copy()
            
            # Try to load from primary dashboard .env file (preferred)
            primary_env_file = '/opt/dashboard-project/es-dashboards/.env'
            local_env_file = '/opt/es-inventory-hub/.env'
            
            # Load environment variables into dict
            env_vars = {}
            for env_file in [primary_env_file, local_env_file]:
                if os.path.exists(env_file):
                    try:
                        with open(env_file, 'r') as f:
                            for line in f:
                                line = line.strip()
                                if line and not line.startswith('#') and '=' in line:
                                    key, value = line.split('=', 1)
                                    env_vars[key.strip()] = value.strip()
                    except Exception as e:
                        pass  # Continue if file can't be read
            
            # Merge loaded env vars into subprocess environment
            env.update(env_vars)
            
            failed_jobs = []
            successful_jobs = []
            
            # Only execute actual collectors (ninja, threatlocker)
            actual_collectors = [c for c in collectors if c in ['ninja', 'threatlocker']]
            
            for collector in actual_collectors:
                try:
                    # Find the job for this collector
                    current_job = None
                    for job in job_runs:
                        if job['job_name'] == f'{collector}-collector':
                            current_job = job
                            break
                    
                    if not current_job:
                        continue
                    
                    # Update job to running
                    from api.progress_tracker import update_job_progress
                    update_job_progress(current_job['job_id'], 'running', 10, f'{collector.title()} collector started')
                    
                    # Execute collector with environment variables
                    if collector == 'ninja':
                        result = subprocess.run([
                            '/opt/es-inventory-hub/.venv/bin/python', '-m', 'collectors.ninja.main'
                        ], cwd='/opt/es-inventory-hub', capture_output=True, text=True, timeout=600, env=env)
                    elif collector == 'threatlocker':
                        result = subprocess.run([
                            '/opt/es-inventory-hub/.venv/bin/python', '-m', 'collectors.threatlocker.main'
                        ], cwd='/opt/es-inventory-hub', capture_output=True, text=True, timeout=600, env=env)
                    else:
                        raise ValueError(f"Unknown collector: {collector}")
                    
                    # Check result and update status
                    if result.returncode == 0:
                        update_job_progress(current_job['job_id'], 'completed', 100, f'{collector.title()} collector completed successfully')
                        successful_jobs.append(f'{collector}-collector')
                    else:
                        error_msg = f'{collector.title()} collector failed: {result.stderr}'
                        update_job_progress(current_job['job_id'], 'failed', 0, error_msg)
                        failed_jobs.append(f'{collector}-collector')
                        
                except subprocess.TimeoutExpired:
                    error_msg = f'{collector.title()} collector timed out after 10 minutes'
                    update_job_progress(current_job['job_id'], 'failed', 0, error_msg)
                    failed_jobs.append(f'{collector}-collector')
                    
                except Exception as e:
                    error_msg = f'Error running {collector} collector: {str(e)}'
                    update_job_progress(current_job['job_id'], 'failed', 0, error_msg)
                    failed_jobs.append(f'{collector}-collector')
            
            # Determine if batch failed (all collectors failed)
            batch_failed = len(successful_jobs) == 0 and len(failed_jobs) > 0
            
            # Execute cross-vendor checks if at least one collector succeeded
            if run_cross_vendor and len(successful_jobs) > 0:
                try:
                    # Find the cross-vendor job
                    cross_vendor_job = None
                    for job in job_runs:
                        if job['job_name'] == 'cross-vendor-checks':
                            cross_vendor_job = job
                            break
                    
                    if cross_vendor_job:
                        # Update job to running
                        from api.progress_tracker import update_job_progress
                        update_job_progress(cross_vendor_job['job_id'], 'running', 10, 'Cross-vendor checks started')
                        
                        # Run cross-vendor checks
                        python_code = """
import sys
sys.path.append('/opt/es-inventory-hub')
from collectors.checks.cross_vendor import run_cross_vendor_checks
from common.db import session_scope
from datetime import date

try:
    with session_scope() as session:
        results = run_cross_vendor_checks(session, date.today())
        print(f'SUCCESS: {results}')
except Exception as e:
    print(f'ERROR: {e}')
    sys.exit(1)
"""
                        
                        result = subprocess.run([
                            '/opt/es-inventory-hub/.venv/bin/python', '-c', python_code
                        ], cwd='/opt/es-inventory-hub', capture_output=True, text=True, timeout=300)
                        
                        if result.returncode == 0:
                            update_job_progress(cross_vendor_job['job_id'], 'completed', 100, 'Cross-vendor checks completed successfully')
                        else:
                            update_job_progress(cross_vendor_job['job_id'], 'failed', 0, f'Cross-vendor checks failed: {result.stderr}')
                            batch_failed = True
                            failed_jobs.append('cross-vendor-checks')
                            
                except Exception as e:
                    # Update job to failed
                    for job in job_runs:
                        if job['job_name'] == 'cross-vendor-checks':
                            from api.progress_tracker import update_job_progress
                            update_job_progress(job['job_id'], 'failed', 0, f'Error starting cross-vendor checks: {str(e)}')
                            batch_failed = True
                            failed_jobs.append('cross-vendor-checks')
                            break
            elif run_cross_vendor and len(successful_jobs) == 0:
                # Mark cross-vendor as cancelled (no collectors succeeded)
                for job in job_runs:
                    if job['job_name'] == 'cross-vendor-checks':
                        from api.progress_tracker import update_job_progress
                        update_job_progress(job['job_id'], 'cancelled', 0, 'Cancelled - no collectors succeeded')
                        break
            
            # Execute Windows 11 24H2 assessment if at least one collector succeeded
            if len(successful_jobs) > 0:
                try:
                    # Find the Windows 11 24H2 assessment job
                    windows_24h2_job = None
                    for job in job_runs:
                        if job['job_name'] == 'windows-11-24h2-assessment':
                            windows_24h2_job = job
                            break
                    
                    if windows_24h2_job:
                        # Update job to running
                        from api.progress_tracker import update_job_progress
                        update_job_progress(windows_24h2_job['job_id'], 'running', 10, 'Windows 11 24H2 assessment started')
                        
                        # Run Windows 11 24H2 assessment
                        result = subprocess.run([
                            '/opt/es-inventory-hub/.venv/bin/python', '/opt/es-inventory-hub/collectors/assessments/windows_11_24h2_assessment.py'
                        ], cwd='/opt/es-inventory-hub', capture_output=True, text=True, timeout=300)
                        
                        if result.returncode == 0:
                            update_job_progress(windows_24h2_job['job_id'], 'completed', 100, 'Windows 11 24H2 assessment completed successfully')
                        else:
                            update_job_progress(windows_24h2_job['job_id'], 'failed', 0, f'Windows 11 24H2 assessment failed: {result.stderr}')
                            batch_failed = True
                            failed_jobs.append('windows-11-24h2-assessment')
                            
                except Exception as e:
                    # Update job to failed
                    for job in job_runs:
                        if job['job_name'] == 'windows-11-24h2-assessment':
                            from api.progress_tracker import update_job_progress
                            update_job_progress(job['job_id'], 'failed', 0, f'Error starting Windows 11 24H2 assessment: {str(e)}')
                            batch_failed = True
                            failed_jobs.append('windows-11-24h2-assessment')
                            break
            else:
                # Mark Windows 11 24H2 assessment as cancelled (no collectors succeeded)
                for job in job_runs:
                    if job['job_name'] == 'windows-11-24h2-assessment':
                        from api.progress_tracker import update_job_progress
                        update_job_progress(job['job_id'], 'cancelled', 0, 'Cancelled - no collectors succeeded')
                        break
            
            # Update batch status based on results
            # Batch is successful if at least one collector succeeded
            batch_status = 'completed' if len(successful_jobs) > 0 else 'failed'
            if len(successful_jobs) > 0 and len(failed_jobs) > 0:
                batch_message = f'Partial success: {len(successful_jobs)} collector(s) succeeded, {len(failed_jobs)} failed: {", ".join(failed_jobs)}'
            elif len(successful_jobs) > 0:
                batch_message = 'All collectors completed successfully'
            else:
                batch_message = f'All collectors failed: {", ".join(failed_jobs)}'
            
            # Update batch status in database
            batch_update_query = text("""
                UPDATE job_batches 
                SET status = :status, message = :message, ended_at = :ended_at
                WHERE batch_id = :batch_id
            """)
            session.execute(batch_update_query, {
                'status': batch_status,
                'message': batch_message,
                'ended_at': datetime.utcnow(),
                'batch_id': batch_id
            })
            session.commit()
            
        return jsonify({
                'batch_id': batch_id,
                'status': batch_status,
                'message': batch_message,
                'failed_jobs': failed_jobs,
                'collectors': job_runs
            }), 201
            
    except Exception as e:
        # If we have a batch_id, mark it as failed
        if 'batch_id' in locals():
            try:
                with get_session() as session:
                    batch_update_query = text("""
                        UPDATE job_batches 
                        SET status = 'failed', message = :message, ended_at = :ended_at
                        WHERE batch_id = :batch_id
                    """)
                    session.execute(batch_update_query, {
                        'message': f'Batch failed due to system error: {str(e)}',
                        'ended_at': datetime.utcnow(),
                        'batch_id': batch_id
                    })
                    session.commit()
            except:
                pass  # Don't fail on cleanup errors
        
        return jsonify({
            'success': False,
            'error': f'System error: {str(e)}',
            'batch_id': batch_id if 'batch_id' in locals() else None
        }), 500

@app.route('/api/collectors/status', methods=['GET'])
def get_collector_status():
    """Get status of collector services."""
    try:
        # Check systemd service status
        ninja_result = subprocess.run([
            'systemctl', 'is-active', 'es-inventory-ninja.service'
        ], capture_output=True, text=True)
        
        threatlocker_result = subprocess.run([
            'systemctl', 'is-active', 'es-inventory-threatlocker.service'
        ], capture_output=True, text=True)
        
        # Get last run times
        ninja_status = subprocess.run([
            'systemctl', 'show', 'es-inventory-ninja.service', '--property=ActiveEnterTimestamp'
        ], capture_output=True, text=True)
        
        threatlocker_status = subprocess.run([
            'systemctl', 'show', 'es-inventory-threatlocker.service', '--property=ActiveEnterTimestamp'
        ], capture_output=True, text=True)
        
        return jsonify({
            'ninja': {
                'status': ninja_result.stdout.strip(),
                'last_run': ninja_status.stdout.strip().split('=')[1] if '=' in ninja_status.stdout else None
            },
            'threatlocker': {
                'status': threatlocker_result.stdout.strip(),
                'last_run': threatlocker_status.stdout.strip().split('=')[1] if '=' in threatlocker_status.stdout else None
            }
        })
        
    except Exception as e:
        return jsonify({
            'error': str(e)
        }), 500

@app.route('/api/collectors/history', methods=['GET'])
def get_collection_history():
    """
    Get collection history for the last 10 runs.
    
    Returns collection runs with timestamps, status, and duration.
    """
    limit = int(request.args.get('limit', 10))
    
    with get_session() as session:
        # Get recent job runs including running jobs
        query = text("""
            SELECT 
                job_id,
                job_name,
                started_at,
                ended_at,
                status,
                message,
                progress_percent,
                updated_at,
                CASE 
                    WHEN ended_at IS NOT NULL THEN 
                        EXTRACT(EPOCH FROM (ended_at - started_at))::INTEGER
                    ELSE NULL
                END as duration_seconds
            FROM job_runs
            WHERE job_name IN ('ninja-collector', 'threatlocker-collector')
            ORDER BY started_at DESC
            LIMIT :limit
        """)
        
        results = session.execute(query, {'limit': limit}).fetchall()
        
        # Format results
        history = []
        for row in results:
            job_id, job_name, started_at, ended_at, status, message, progress_percent, updated_at, duration_seconds = row
            
            # Calculate duration
            duration_str = None
            if duration_seconds is not None:
                if duration_seconds < 60:
                    duration_str = f"{duration_seconds}s"
                elif duration_seconds < 3600:
                    minutes = duration_seconds // 60
                    seconds = duration_seconds % 60
                    duration_str = f"{minutes}m {seconds}s"
                else:
                    hours = duration_seconds // 3600
                    minutes = (duration_seconds % 3600) // 60
                    duration_str = f"{hours}h {minutes}m"
            
            history.append({
                'job_id': job_id,
                'job_name': job_name,
                'started_at': started_at.isoformat() + 'Z' if started_at else None,
                'ended_at': ended_at.isoformat() + 'Z' if ended_at else None,
                'status': status,
                'message': message,
                'progress_percent': progress_percent,
                'updated_at': updated_at.isoformat() + 'Z' if updated_at else None,
                'duration': duration_str,
                'duration_seconds': duration_seconds
            })
        
        return jsonify({
            'collection_history': history,
            'total_runs': len(history),
            'generated_at': datetime.now().isoformat()
        })

def cleanup_stale_jobs(session):
    """
    Detect and clean up stale running jobs that have no active process.
    Returns list of cleaned up job IDs.
    """
    cleaned_jobs = []
    
    # Find jobs that have been "running" for more than 10 minutes
    stale_query = text("""
        SELECT 
            job_id,
            job_name,
            started_at,
            EXTRACT(EPOCH FROM (NOW() - started_at))::int as seconds_running
        FROM job_runs
        WHERE status IN ('running', 'queued')
        AND started_at < NOW() - INTERVAL '10 minutes'
        AND job_name IN ('ninja-collector', 'threatlocker-collector')
    """)
    
    stale_jobs = session.execute(stale_query).fetchall()
    
    for job_id, job_name, started_at, seconds_running in stale_jobs:
        # Check if process is actually running
        process_running = False
        
        # Check for Python collector process
        try:
            if job_name == 'ninja-collector':
                proc_check = subprocess.run([
                    'pgrep', '-f', 'collectors.ninja.main'
                ], capture_output=True, text=True)
            elif job_name == 'threatlocker-collector':
                proc_check = subprocess.run([
                    'pgrep', '-f', 'collectors.threatlocker.main'
                ], capture_output=True, text=True)
            else:
                proc_check = None
            
            if proc_check and proc_check.returncode == 0:
                process_running = True
        except:
            pass
        
        # If no process is running, mark job as failed
        if not process_running:
            update_query = text("""
                UPDATE job_runs
                SET status = 'failed',
                    message = 'Job appears to have failed or was interrupted - no process found running',
                    ended_at = NOW(),
                    updated_at = NOW()
                WHERE job_id = :job_id
            """)
            session.execute(update_query, {'job_id': job_id})
            cleaned_jobs.append({
                'job_id': job_id,
                'job_name': job_name,
                'started_at': started_at.isoformat() if started_at else None,
                'seconds_running': seconds_running,
                'reason': 'No active process found'
            })
    
    if cleaned_jobs:
        session.commit()
    
    return cleaned_jobs

@app.route('/api/collectors/progress', methods=['GET'])
def get_collection_progress():
    """
    Get real-time collection progress if collectors are currently running.
    Automatically cleans up stale running jobs.
    
    Returns progress information for active collection jobs.
    """
    with get_session() as session:
        # Clean up stale jobs first
        cleaned_jobs = cleanup_stale_jobs(session)
        
        # Check for active collections
        query = text("""
            SELECT 
                job_id,
                job_name,
                started_at,
                status,
                message
            FROM job_runs
            WHERE job_name IN ('ninja-collector', 'threatlocker-collector')
            AND status IN ('running', 'queued')
            AND started_at >= CURRENT_DATE
            ORDER BY started_at DESC
        """)
        
        results = session.execute(query).fetchall()
        
        # Check systemd service status for additional info
        active_collections = []
        for row in results:
            job_id, job_name, started_at, status, message = row
            
            # Check if process is actually running
            process_running = False
            try:
                if job_name == 'ninja-collector':
                    proc_check = subprocess.run([
                        'pgrep', '-f', 'collectors.ninja.main'
                    ], capture_output=True, text=True)
                elif job_name == 'threatlocker-collector':
                    proc_check = subprocess.run([
                        'pgrep', '-f', 'collectors.threatlocker.main'
                    ], capture_output=True, text=True)
                else:
                    proc_check = None
                
                if proc_check and proc_check.returncode == 0:
                    process_running = True
            except:
                pass
            
            # Get systemd status
            service_name = f"{job_name}.service"
            if job_name == 'threatlocker-collector':
                service_name = "threatlocker-collector@rene.service"
            
            try:
                systemd_result = subprocess.run([
                    'systemctl', 'is-active', service_name
                ], capture_output=True, text=True)
                
                is_active = systemd_result.stdout.strip() == 'active'
            except:
                is_active = False
            
            # Calculate elapsed time
            elapsed_seconds = None
            if started_at:
                elapsed = datetime.now() - started_at
                elapsed_seconds = int(elapsed.total_seconds())
            
            active_collections.append({
                'job_id': job_id,
                'job_name': job_name,
                'started_at': started_at.isoformat() if started_at else None,
                'status': status,
                'message': message,
                'is_active': is_active,
                'process_running': process_running,
                'elapsed_seconds': elapsed_seconds,
                'elapsed_time': f"{elapsed_seconds // 60}m {elapsed_seconds % 60}s" if elapsed_seconds else None
            })
        
        return jsonify({
            'active_collections': active_collections,
            'total_active': len(active_collections),
            'cleaned_stale_jobs': cleaned_jobs,
            'generated_at': datetime.now().isoformat()
        })

@app.route('/api/collectors/cleanup-stale', methods=['POST'])
def cleanup_stale_jobs_endpoint():
    """
    Manually trigger cleanup of stale running jobs.
    
    This endpoint allows the dashboard to proactively clean up jobs that appear
    to be running but have no active process. Useful when dashboard detects
    jobs that have been "running" for an unusually long time.
    
    Returns list of cleaned up jobs.
    """
    try:
        with get_session() as session:
            cleaned_jobs = cleanup_stale_jobs(session)
            
            return jsonify({
                'success': True,
                'cleaned_jobs': cleaned_jobs,
                'total_cleaned': len(cleaned_jobs),
                'message': f'Cleaned up {len(cleaned_jobs)} stale job(s)' if cleaned_jobs else 'No stale jobs found'
            })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/collectors/runs/batch/<batch_id>', methods=['GET', 'OPTIONS'])
def get_batch_status(batch_id):
    """Get status of a specific batch run."""
    try:
        with get_session() as session:
            # Get batch info
            batch_query = text("""
                SELECT batch_id, created_at, status, priority, started_at, ended_at, 
                       progress_percent, estimated_completion, message, error, duration_seconds
                FROM job_batches 
                WHERE batch_id = :batch_id
            """)
            batch_result = session.execute(batch_query, {'batch_id': batch_id}).fetchone()
            
            if not batch_result:
                return jsonify({'error': 'Batch not found'}), 404
            
            # Get job runs for this batch
            jobs_query = text("""
                SELECT job_id, job_name, status, started_at, updated_at, ended_at,
                       progress_percent, message, error, duration_seconds
                FROM job_runs 
                WHERE batch_id = :batch_id
                ORDER BY started_at
            """)
            jobs_results = session.execute(jobs_query, {'batch_id': batch_id}).fetchall()
            
            # Format response
            collectors = []
            for job in jobs_results:
                collectors.append({
                    'job_name': job.job_name,
                    'job_id': job.job_id,
                    'status': job.status,
                    'started_at': job.started_at.isoformat() + 'Z' if job.started_at else None,
                    'updated_at': job.updated_at.isoformat() + 'Z' if job.updated_at else None,
                    'ended_at': job.ended_at.isoformat() + 'Z' if job.ended_at else None,
                    'progress_percent': job.progress_percent,
                    'message': job.message,
                    'error': job.error,
                    'duration_seconds': job.duration_seconds
                })
            
            return jsonify({
                'batch_id': batch_result.batch_id,
                'status': batch_result.status,
                'progress_percent': batch_result.progress_percent,
                'estimated_completion': batch_result.estimated_completion.isoformat() + 'Z' if batch_result.estimated_completion else None,
                'started_at': batch_result.started_at.isoformat() + 'Z' if batch_result.started_at else None,
                'updated_at': datetime.utcnow().isoformat() + 'Z',
                'ended_at': batch_result.ended_at.isoformat() + 'Z' if batch_result.ended_at else None,
                'message': batch_result.message,
                'error': batch_result.error,
                'collectors': collectors,
                'duration_seconds': batch_result.duration_seconds
            })
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/collectors/runs/latest', methods=['GET', 'OPTIONS'])
def get_latest_runs():
    """Get latest active or most recent terminal runs for specified collectors."""
    collectors_param = request.args.get('collectors', 'ninja,threatlocker')
    collectors = [c.strip() for c in collectors_param.split(',')]
    
    try:
        with get_session() as session:
            # Get latest runs for each collector
            latest_runs = []
            for collector in collectors:
                job_name = f"{collector}-collector"
                
                # First try to get active runs
                active_query = text("""
                    SELECT jr.job_id, jr.batch_id, jr.job_name, jr.status, jr.started_at,
                           jr.updated_at, jr.ended_at, jr.progress_percent, jr.message,
                           jr.error, jr.duration_seconds,
                           jb.status as batch_status, jb.progress_percent as batch_progress,
                           jb.estimated_completion, jb.message as batch_message
                    FROM job_runs jr
                    JOIN job_batches jb ON jr.batch_id = jb.batch_id
                    WHERE jr.job_name = :job_name AND jr.status IN ('queued', 'running')
                    ORDER BY jr.started_at DESC
                    LIMIT 1
                """)
                active_result = session.execute(active_query, {'job_name': job_name}).fetchone()
                
                if active_result:
                    latest_runs.append({
                        'job_id': active_result.job_id,
                        'batch_id': active_result.batch_id,
                        'job_name': active_result.job_name,
                        'status': active_result.status,
                        'started_at': active_result.started_at.isoformat() + 'Z' if active_result.started_at else None,
                        'updated_at': active_result.updated_at.isoformat() + 'Z' if active_result.updated_at else None,
                        'ended_at': active_result.ended_at.isoformat() + 'Z' if active_result.ended_at else None,
                        'progress_percent': active_result.progress_percent,
                        'message': active_result.message,
                        'error': active_result.error,
                        'duration_seconds': active_result.duration_seconds,
                        'batch_status': active_result.batch_status,
                        'batch_progress_percent': active_result.batch_progress,
                        'estimated_completion': active_result.estimated_completion.isoformat() + 'Z' if active_result.estimated_completion else None,
                        'batch_message': active_result.batch_message
                    })
                else:
                    # Get most recent terminal run
                    terminal_query = text("""
                        SELECT jr.job_id, jr.batch_id, jr.job_name, jr.status, jr.started_at,
                               jr.updated_at, jr.ended_at, jr.progress_percent, jr.message,
                               jr.error, jr.duration_seconds,
                               jb.status as batch_status, jb.progress_percent as batch_progress,
                               jb.estimated_completion, jb.message as batch_message
                        FROM job_runs jr
                        JOIN job_batches jb ON jr.batch_id = jb.batch_id
                        WHERE jr.job_name = :job_name AND jr.status IN ('completed', 'failed', 'cancelled')
                        ORDER BY jr.started_at DESC
                        LIMIT 1
                    """)
                    terminal_result = session.execute(terminal_query, {'job_name': job_name}).fetchone()
                    
                    if terminal_result:
                        latest_runs.append({
                            'job_id': terminal_result.job_id,
                            'batch_id': terminal_result.batch_id,
                            'job_name': terminal_result.job_name,
                            'status': terminal_result.status,
                            'started_at': terminal_result.started_at.isoformat() + 'Z' if terminal_result.started_at else None,
                            'updated_at': terminal_result.updated_at.isoformat() + 'Z' if terminal_result.updated_at else None,
                            'ended_at': terminal_result.ended_at.isoformat() + 'Z' if terminal_result.ended_at else None,
                            'progress_percent': terminal_result.progress_percent,
                            'message': terminal_result.message,
                            'error': terminal_result.error,
                            'duration_seconds': terminal_result.duration_seconds,
                            'batch_status': terminal_result.batch_status,
                            'batch_progress_percent': terminal_result.batch_progress,
                            'estimated_completion': terminal_result.estimated_completion.isoformat() + 'Z' if terminal_result.estimated_completion else None,
                            'batch_message': terminal_result.batch_message
                        })
            
            return jsonify({
                'latest_runs': latest_runs,
                'total_runs': len(latest_runs),
                'generated_at': datetime.utcnow().isoformat() + 'Z'
            })
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/collectors/runs/job/<job_id>', methods=['GET', 'OPTIONS'])
def get_job_status(job_id):
    """Get status of a specific job run."""
    try:
        with get_session() as session:
            # Get job info
            job_query = text("""
                SELECT jr.job_id, jr.batch_id, jr.job_name, jr.status, jr.started_at, 
                       jr.updated_at, jr.ended_at, jr.progress_percent, jr.message, 
                       jr.error, jr.duration_seconds,
                       jb.status as batch_status, jb.progress_percent as batch_progress,
                       jb.estimated_completion, jb.message as batch_message
                FROM job_runs jr
                JOIN job_batches jb ON jr.batch_id = jb.batch_id
                WHERE jr.job_id = :job_id
            """)
            job_result = session.execute(job_query, {'job_id': job_id}).fetchone()
            
            if not job_result:
                return jsonify({'error': 'Job not found'}), 404
            
            return jsonify({
                'job_id': job_result.job_id,
                'batch_id': job_result.batch_id,
                'job_name': job_result.job_name,
                'status': job_result.status,
                'started_at': job_result.started_at.isoformat() + 'Z' if job_result.started_at else None,
                'updated_at': job_result.updated_at.isoformat() + 'Z' if job_result.updated_at else None,
                'ended_at': job_result.ended_at.isoformat() + 'Z' if job_result.ended_at else None,
                'progress_percent': job_result.progress_percent,
                'message': job_result.message,
                'error': job_result.error,
                'duration_seconds': job_result.duration_seconds,
                'batch_status': job_result.batch_status,
                'batch_progress_percent': job_result.batch_progress,
                'estimated_completion': job_result.estimated_completion.isoformat() + 'Z' if job_result.estimated_completion else None,
                'batch_message': job_result.batch_message
            })
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/docs/dashboard-ai-guide', methods=['GET'])
def get_dashboard_ai_guide():
    """Serve Dashboard AI integration guide."""
    try:
        with open('/opt/es-inventory-hub/docs/DASHBOARD_AI_COLLECTOR_TRACKING_GUIDE.md', 'r') as f:
            content = f.read()
        return content, 200, {'Content-Type': 'text/markdown; charset=utf-8'}
    except FileNotFoundError:
        return "Documentation not found", 404
    except Exception as e:
        return f"Error reading documentation: {str(e)}", 500

@app.route('/api/docs/api-integration', methods=['GET'])
def get_api_integration():
    """Serve Dashboard AI integration documentation (API_INTEGRATION.md)."""
    try:
        with open('/opt/es-inventory-hub/docs/API_INTEGRATION.md', 'r') as f:
            content = f.read()
        return content, 200, {'Content-Type': 'text/markdown; charset=utf-8'}
    except FileNotFoundError:
        return "Documentation not found", 404
    except Exception as e:
        return f"Error reading documentation: {str(e)}", 500

@app.route('/api/docs/collector-tracking-api', methods=['GET'])
def get_collector_tracking_api():
    """Serve collector tracking API documentation."""
    try:
        with open('/opt/es-inventory-hub/docs/COLLECTOR_RUN_TRACKING_API.md', 'r') as f:
            content = f.read()
        return content, 200, {'Content-Type': 'text/markdown; charset=utf-8'}
    except FileNotFoundError:
        return "Documentation not found", 404
    except Exception as e:
        return f"Error reading documentation: {str(e)}", 500

@app.route('/api/docs', methods=['GET'])
def get_docs_index():
    """Serve documentation index."""
    docs = {
        "dashboard_ai_guide": "/api/docs/dashboard-ai-guide",
        "api_integration": "/api/docs/api-integration",
        "collector_tracking_api": "/api/docs/collector-tracking-api"
    }
    return jsonify({
        "message": "ES Inventory Hub API Documentation",
        "available_docs": docs,
        "base_url": "https://db-api.enersystems.com:5400"
        })

@app.route('/api/exceptions', methods=['GET'])
def get_exceptions():
    """Get exceptions with filtering options."""
    # Parse query parameters
    exception_type = request.args.get('type')
    resolved = request.args.get('resolved')
    limit = int(request.args.get('limit', 100))
    offset = int(request.args.get('offset', 0))
    
    with get_session() as session:
        query = text("""
            SELECT id, date_found, type, hostname, details, resolved
            FROM exceptions
            WHERE 1=1
        """)
        
        params = {}
        
        if exception_type:
            query += " AND type = :exception_type"
            params['exception_type'] = exception_type
        
        if resolved is not None:
            query += " AND resolved = :resolved"
            params['resolved'] = resolved.lower() == 'true'
        
        query += " ORDER BY date_found DESC, type, hostname"
        query += " LIMIT :limit OFFSET :offset"
        params['limit'] = limit
        params['offset'] = offset
        
        exceptions = session.execute(query, params).fetchall()
        
        return jsonify([{
            'id': exc[0],
            'date_found': exc[1].isoformat(),
            'type': exc[2],
            'hostname': exc[3],
            'details': exc[4],
            'resolved': exc[5]
        } for exc in exceptions])

@app.route('/api/exceptions/<int:exception_id>/resolve', methods=['POST'])
def resolve_exception(exception_id: int):
    """Mark an exception as resolved."""
    data = request.get_json() or {}
    resolved_by = data.get('resolved_by', 'api_user')
    notes = data.get('notes', '')
    
    with get_session() as session:
        query = text("""
            UPDATE exceptions
            SET resolved = TRUE, resolved_date = CURRENT_DATE, resolved_by = :resolved_by
            WHERE id = :exception_id
        """)
        
        result = session.execute(query, {
            'exception_id': exception_id,
            'resolved_by': resolved_by
        })
        
        session.commit()
        
        if result.rowcount == 0:
            return jsonify({'error': 'Exception not found'}), 404
        
        return jsonify({'success': True, 'message': 'Exception resolved'})

@app.route('/api/exceptions/<int:exception_id>/mark-manually-fixed', methods=['POST'])
def mark_exception_manually_fixed(exception_id: int):
    """
    Mark an exception as manually fixed by dashboard user.
    
    This endpoint addresses the critical gap where dashboard updates
    ThreatLocker but the database doesn't know about manual fixes.
    """
    data = request.get_json() or {}
    updated_by = data.get('updated_by', 'dashboard_user')
    update_type = data.get('update_type', 'unknown')
    old_value = data.get('old_value', {})
    new_value = data.get('new_value', {})
    notes = data.get('notes', '')
    
    with get_session() as session:
        # First check if exception exists
        check_query = text("SELECT id, hostname, type FROM exceptions WHERE id = :exception_id")
        result = session.execute(check_query, {'exception_id': exception_id}).fetchone()
        
        if not result:
            return jsonify({'error': 'Exception not found'}), 404
        
        # Update exception with manual fix information
        update_query = text("""
            UPDATE exceptions
            SET 
                resolved = TRUE,
                manually_updated_at = CURRENT_TIMESTAMP,
                manually_updated_by = :updated_by,
                update_type = :update_type,
                old_value = :old_value,
                new_value = :new_value,
                variance_status = 'manually_fixed'
            WHERE id = :exception_id
        """)
        
        session.execute(update_query, {
            'exception_id': exception_id,
            'updated_by': updated_by,
            'update_type': update_type,
            'old_value': json.dumps(old_value),
            'new_value': json.dumps(new_value)
        })
        
        session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Exception marked as manually fixed',
            'exception_id': exception_id,
            'hostname': result[1],
            'type': result[2],
            'updated_by': updated_by,
            'updated_at': datetime.now().isoformat()
        })

@app.route('/api/exceptions/mark-fixed-by-hostname', methods=['POST'])
def mark_exception_fixed_by_hostname():
    """
    Mark exceptions as manually fixed by hostname and exception type.
    
    This endpoint allows the dashboard to mark exceptions as fixed when
    they update ThreatLocker display names, without needing to know
    the specific exception ID.
    
    FIXED: Now handles both unresolved and already-resolved exceptions
    to provide proper feedback to the dashboard.
    """
    data = request.get_json() or {}
    hostname = data.get('hostname')
    exception_type = data.get('type', 'DISPLAY_NAME_MISMATCH')
    updated_by = data.get('updated_by', 'dashboard_user')
    update_type = data.get('update_type', 'display_name_update')
    old_value = data.get('old_value', {})
    new_value = data.get('new_value', {})
    notes = data.get('notes', '')
    
    if not hostname:
        return jsonify({'error': 'Hostname is required'}), 400
    
    with get_session() as session:
        # First, check if ANY exceptions exist for this hostname (resolved or unresolved)
        check_query = text("""
            SELECT id, hostname, type, resolved, manually_updated_at, manually_updated_by
            FROM exceptions 
            WHERE hostname = :hostname 
            AND type = :exception_type
            AND date_found = CURRENT_DATE
            ORDER BY id DESC
        """)
        
        all_exceptions = session.execute(check_query, {
            'hostname': hostname,
            'exception_type': exception_type
        }).fetchall()
        
        if not all_exceptions:
            return jsonify({
                'success': False,
                'message': f'No {exception_type} exceptions found for hostname {hostname}',
                'hostname': hostname,
                'type': exception_type,
                'status': 'not_found'
            }), 404
        
        # Check if any are unresolved
        unresolved_exceptions = [exc for exc in all_exceptions if not exc[3]]  # resolved is index 3
        
        if unresolved_exceptions:
            # Update unresolved exceptions
            update_query = text("""
                UPDATE exceptions
                SET 
                    resolved = TRUE,
                    manually_updated_at = CURRENT_TIMESTAMP,
                    manually_updated_by = :updated_by,
                    update_type = :update_type,
                    old_value = :old_value,
                    new_value = :new_value,
                    variance_status = 'manually_fixed'
                WHERE hostname = :hostname 
                AND type = :exception_type
                AND resolved = FALSE
                AND date_found = CURRENT_DATE
            """)
            
            result = session.execute(update_query, {
                'hostname': hostname,
                'exception_type': exception_type,
                'updated_by': updated_by,
                'update_type': update_type,
                'old_value': json.dumps(old_value),
                'new_value': json.dumps(new_value)
            })
            
            session.commit()
            
            return jsonify({
                'success': True,
                'message': f'Marked {result.rowcount} exceptions as manually fixed',
                'hostname': hostname,
                'type': exception_type,
                'exceptions_updated': result.rowcount,
                'updated_by': updated_by,
                'updated_at': datetime.now().isoformat(),
                'status': 'updated'
            })
        else:
            # All exceptions are already resolved - return success with info
            latest_exception = all_exceptions[0]  # Most recent
            return jsonify({
                'success': True,
                'message': f'Exception already resolved for hostname {hostname}',
                'hostname': hostname,
                'type': exception_type,
                'exceptions_updated': 0,
                'status': 'already_resolved',
                'last_updated_by': latest_exception[5],  # manually_updated_by
                'last_updated_at': latest_exception[4].isoformat() if latest_exception[4] else None,  # manually_updated_at
                'updated_at': datetime.now().isoformat()
            })

@app.route('/api/exceptions/bulk-update', methods=['POST'])
def bulk_update_exceptions():
    """
    Perform bulk operations on multiple exceptions.
    
    Supports bulk marking as manually fixed, resolved, or status changes.
    """
    data = request.get_json() or {}
    exception_ids = data.get('exception_ids', [])
    action = data.get('action', 'mark_manually_fixed')
    updated_by = data.get('updated_by', 'dashboard_user')
    notes = data.get('notes', '')
    
    if not exception_ids:
        return jsonify({'error': 'No exception IDs provided'}), 400
    
    if action not in ['mark_manually_fixed', 'resolve', 'reset_status']:
        return jsonify({'error': 'Invalid action'}), 400
    
    with get_session() as session:
        # Build dynamic query based on action
        if action == 'mark_manually_fixed':
            query = text("""
                UPDATE exceptions
                SET 
                    resolved = TRUE,
                    manually_updated_at = CURRENT_TIMESTAMP,
                    manually_updated_by = :updated_by,
                    variance_status = 'manually_fixed'
                WHERE id = ANY(:exception_ids)
            """)
        elif action == 'resolve':
            query = text("""
                UPDATE exceptions
                SET 
                    resolved = TRUE,
                    resolved_date = CURRENT_DATE,
                    resolved_by = :updated_by
                WHERE id = ANY(:exception_ids)
            """)
        elif action == 'reset_status':
            query = text("""
                UPDATE exceptions
                SET 
                    variance_status = 'active',
                    manually_updated_at = NULL,
                    manually_updated_by = NULL
                WHERE id = ANY(:exception_ids)
            """)
        
        result = session.execute(query, {
            'exception_ids': exception_ids,
            'updated_by': updated_by
        })
        
        session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Bulk {action} completed',
            'updated_count': result.rowcount,
            'exception_ids': exception_ids,
            'updated_by': updated_by
        })

@app.route('/api/collectors/threatlocker/run', methods=['POST'])
def run_threatlocker_collector():
    """
    Run the ThreatLocker collector script via API.
    
    This endpoint allows the dashboard to trigger the ThreatLocker collector
    to refresh all device data from the ThreatLocker API.
    """
    data = request.get_json() or {}
    force_refresh = data.get('force_refresh', False)
    organization_id = data.get('organization_id', None)
    
    try:
        import subprocess
        import sys
        import os
        
        # Change to the project directory
        project_dir = '/opt/es-inventory-hub'
        
        # Build the command
        cmd = [sys.executable, '-m', 'collectors.threatlocker.main']
        
        # Add any additional parameters if needed
        if organization_id:
            cmd.extend(['--organization', organization_id])
        
        # Run the collector
        result = subprocess.run(
            cmd,
            cwd=project_dir,
            capture_output=True,
            text=True,
            timeout=300  # 5 minute timeout
        )
        
        if result.returncode == 0:
            return jsonify({
                'success': True,
                'message': 'ThreatLocker collector completed successfully',
                'output': result.stdout,
                'force_refresh': force_refresh,
                'organization_id': organization_id
            })
        else:
            return jsonify({
                'success': False,
                'message': 'ThreatLocker collector failed',
                'error': result.stderr,
                'return_code': result.returncode
            }), 500
            
    except subprocess.TimeoutExpired:
        return jsonify({
            'success': False,
            'message': 'ThreatLocker collector timed out after 5 minutes'
        }), 500
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Failed to run ThreatLocker collector: {str(e)}'
        }), 500

@app.route('/api/collectors/cross-vendor/run', methods=['POST'])
def run_cross_vendor_checks():
    """
    Run the cross-vendor consistency checks via API.
    
    This endpoint allows the dashboard to trigger cross-vendor checks
    to update variance data after collector runs.
    """
    data = request.get_json() or {}
    force_refresh = data.get('force_refresh', False)
    
    try:
        import sys
        import os
        
        # Change to the project directory
        project_dir = '/opt/es-inventory-hub'
        
        # Build the Python command to run cross-vendor checks
        python_code = """
import sys
sys.path.append('/opt/es-inventory-hub')
from collectors.checks.cross_vendor import run_cross_vendor_checks
from common.db import session_scope
from datetime import date

try:
    with session_scope() as session:
        results = run_cross_vendor_checks(session, date.today())
        print(f'SUCCESS: {results}')
except Exception as e:
    print(f'ERROR: {e}')
    import traceback
    traceback.print_exc()
    sys.exit(1)
"""
        
        # Run the cross-vendor checks
        result = subprocess.run(
            [sys.executable, '-c', python_code],
            cwd=project_dir,
            capture_output=True,
            text=True,
            timeout=120  # 2 minute timeout
        )
        
        if result.returncode == 0:
            return jsonify({
                'success': True,
                'message': 'Cross-vendor checks completed successfully',
                'output': result.stdout,
                'force_refresh': force_refresh
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Cross-vendor checks failed',
                'error': result.stderr,
                'return_code': result.returncode
            }), 500
            
    except subprocess.TimeoutExpired:
        return jsonify({
            'success': False,
            'message': 'Cross-vendor checks timed out after 2 minutes'
        }), 500
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Failed to run cross-vendor checks: {str(e)}'
        }), 500

@app.route('/api/collectors/sequence/run', methods=['POST'])
def run_collector_sequence():
    """
    Run the complete collector sequence in proper order via API.
    
    This endpoint runs: ThreatLocker collector → cross-vendor checks
    to ensure complete data refresh and variance update.
    """
    data = request.get_json() or {}
    force_refresh = data.get('force_refresh', False)
    organization_id = data.get('organization_id', None)
    
    try:
        import subprocess
        import sys
        import os
        
        # Change to the project directory
        project_dir = '/opt/es-inventory-hub'
        results = []
        
        # Step 1: Run ThreatLocker collector
        cmd1 = [sys.executable, '-m', 'collectors.threatlocker.main']
        if organization_id:
            cmd1.extend(['--organization', organization_id])
        
        result1 = subprocess.run(
            cmd1,
            cwd=project_dir,
            capture_output=True,
            text=True,
            timeout=300
        )
        
        if result1.returncode != 0:
            return jsonify({
                'success': False,
                'message': 'ThreatLocker collector failed in sequence',
                'step': 'threatlocker_collector',
                'error': result1.stderr,
                'return_code': result1.returncode
            }), 500
        
        results.append({
            'step': 'threatlocker_collector',
            'success': True,
            'output': result1.stdout
        })
        
        # Step 2: Run cross-vendor checks
        python_code = """
import sys
sys.path.append('/opt/es-inventory-hub')
from collectors.checks.cross_vendor import run_cross_vendor_checks
from common.db import session_scope
from datetime import date

try:
    with session_scope() as session:
        results = run_cross_vendor_checks(session, date.today())
        print(f'SUCCESS: {results}')
except Exception as e:
    print(f'ERROR: {e}')
    import traceback
    traceback.print_exc()
    sys.exit(1)
"""
        
        result2 = subprocess.run(
            [sys.executable, '-c', python_code],
            cwd=project_dir,
            capture_output=True,
            text=True,
            timeout=120
        )
        
        if result2.returncode != 0:
            return jsonify({
                'success': False,
                'message': 'Cross-vendor checks failed in sequence',
                'step': 'cross_vendor_checks',
                'error': result2.stderr,
                'return_code': result2.returncode,
                'completed_steps': results
            }), 500
        
        results.append({
            'step': 'cross_vendor_checks',
            'success': True,
            'output': result2.stdout
        })
        
        return jsonify({
            'success': True,
            'message': 'Complete collector sequence completed successfully',
            'steps': results,
            'force_refresh': force_refresh,
            'organization_id': organization_id
        })
        
    except subprocess.TimeoutExpired as e:
        return jsonify({
            'success': False,
            'message': f'Collector sequence timed out: {str(e)}',
            'completed_steps': results
        }), 500
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Failed to run collector sequence: {str(e)}',
            'completed_steps': results
        }), 500

@app.route('/api/threatlocker/update-name', methods=['POST'])
def update_threatlocker_name():
    """
    Update a ThreatLocker computer name via API.
    
    This endpoint allows the dashboard to update ThreatLocker computer names
    and then sync the changes to the ES Inventory Hub database.
    """
    data = request.get_json() or {}
    computer_id = data.get('computer_id', '').strip()
    hostname = data.get('hostname', '').strip()
    new_name = data.get('new_name', '').strip()
    updated_by = data.get('updated_by', 'dashboard_user')
    
    if not new_name:
        return jsonify({'error': 'new_name is required'}), 400
    
    if not computer_id and not hostname:
        return jsonify({'error': 'Either computer_id or hostname is required'}), 400
    
    try:
        # Import ThreatLocker API
        from collectors.threatlocker.api import ThreatLockerAPI
        from common.db import session_scope
        from datetime import date
        
        # Initialize ThreatLocker API
        api = ThreatLockerAPI()
        
        # If only hostname provided, find the computer_id
        if not computer_id and hostname:
            devices = api.fetch_devices()
            target_device = None
            
            for device in devices:
                if device.get('hostname', '').lower() == hostname.lower():
                    target_device = device
                    break
            
            if not target_device:
                return jsonify({
                    'error': f'Device not found in ThreatLocker API',
                    'hostname': hostname
                }), 404
            
            computer_id = target_device.get('computerId')
            if not computer_id:
                return jsonify({
                    'error': f'Device found but missing computerId',
                    'hostname': hostname
                }), 500
        
        # Update the computer name in ThreatLocker
        update_result = api.update_computer_name(computer_id, new_name)
        
        if not update_result['success']:
            return jsonify({
                'error': f'Failed to update ThreatLocker computer name: {update_result.get("error", "Unknown error")}',
                'computer_id': computer_id,
                'new_name': new_name
            }), 500
        
        # Now sync the updated device to the database
        with session_scope() as session:
            # Get vendor ID for ThreatLocker
            vendor_query = text("SELECT id FROM vendor WHERE name = 'ThreatLocker'")
            vendor_result = session.execute(vendor_query).fetchone()
            if not vendor_result:
                return jsonify({'error': 'ThreatLocker vendor not found in database'}), 500
            vendor_id = vendor_result[0]
            
            # Fetch the updated device from ThreatLocker API
            devices = api.fetch_devices()
            updated_device = None
            
            for device in devices:
                if device.get('computerId') == computer_id:
                    updated_device = device
                    break
            
            if not updated_device:
                return jsonify({
                    'error': f'Updated device not found in ThreatLocker API',
                    'computer_id': computer_id
                }), 500
            
            # Import mapping function
            from collectors.threatlocker.mapping import normalize_threatlocker_device
            from common.util import upsert_device_identity, insert_snapshot
            
            # Normalize the updated device data
            normalized = normalize_threatlocker_device(updated_device)
            
            # Get today's date for snapshot
            snapshot_date = date.today()
            
            # Delete existing snapshots for this device on today's date
            delete_query = text("""
                DELETE FROM device_snapshot 
                WHERE snapshot_date = :snapshot_date 
                AND vendor_id = :vendor_id 
                AND device_identity_id IN (
                    SELECT id FROM device_identity 
                    WHERE vendor_id = :vendor_id 
                    AND vendor_device_key = :vendor_device_key
                )
            """)
            
            session.execute(delete_query, {
                'snapshot_date': snapshot_date,
                'vendor_id': vendor_id,
                'vendor_device_key': normalized['vendor_device_key']
            })
            
            # Create or update device identity
            device_identity_id = upsert_device_identity(
                session=session,
                vendor_id=vendor_id,
                vendor_device_key=normalized['vendor_device_key'],
                first_seen_date=snapshot_date
            )
            
            # Insert the updated snapshot
            insert_snapshot(
                session=session,
                snapshot_date=snapshot_date,
                vendor_id=vendor_id,
                device_identity_id=device_identity_id,
                normalized=normalized
            )
            
            session.commit()
            
            return jsonify({
                'success': True,
                'message': 'ThreatLocker computer name updated and synced successfully',
                'device': {
                    'computer_id': computer_id,
                    'hostname': normalized.get('hostname'),
                    'old_name': data.get('old_name', 'Unknown'),
                    'new_name': new_name,
                    'display_name': normalized.get('display_name'),
                    'organization_name': normalized.get('organization_name')
                },
                'threatlocker_update': update_result,
                'updated_by': updated_by,
                'updated_at': datetime.now().isoformat()
            })
            
    except Exception as e:
        return jsonify({
            'error': f'Failed to update ThreatLocker computer name: {str(e)}'
        }), 500

@app.route('/api/threatlocker/sync-device', methods=['POST'])
def sync_threatlocker_device():
    """
    Sync a specific ThreatLocker device to update the database with latest information.
    
    This endpoint allows the dashboard to trigger a database update for a specific
    ThreatLocker device after making changes via the ThreatLocker API.
    
    This solves the critical issue where dashboard updates ThreatLocker names
    but the ES Inventory Hub database doesn't reflect the changes until the next
    scheduled collector run.
    """
    data = request.get_json() or {}
    computer_id = data.get('computer_id', '').strip()
    hostname = data.get('hostname', '').strip()
    updated_by = data.get('updated_by', 'dashboard_user')
    
    if not computer_id and not hostname:
        return jsonify({'error': 'Either computer_id or hostname is required'}), 400
    
    try:
        # Import ThreatLocker API and mapping functions
        from collectors.threatlocker.api import ThreatLockerAPI
        from collectors.threatlocker.mapping import normalize_threatlocker_device
        from common.util import upsert_device_identity, insert_snapshot
        from common.db import session_scope
        from datetime import date
        
        with session_scope() as session:
            # Get vendor ID for ThreatLocker
            vendor_query = text("SELECT id FROM vendor WHERE name = 'ThreatLocker'")
            vendor_result = session.execute(vendor_query).fetchone()
            if not vendor_result:
                return jsonify({'error': 'ThreatLocker vendor not found in database'}), 500
            vendor_id = vendor_result[0]
            
            # Initialize ThreatLocker API
            api = ThreatLockerAPI()
            
            # Fetch the specific device from ThreatLocker API
            devices = api.fetch_devices()
            target_device = None
            
            if computer_id:
                # Find by computerId (preferred method)
                for device in devices:
                    if device.get('computerId') == computer_id:
                        target_device = device
                        break
            else:
                # Find by hostname (fallback)
                for device in devices:
                    if device.get('hostname', '').lower() == hostname.lower():
                        target_device = device
                        break
            
            if not target_device:
                return jsonify({
                    'error': f'Device not found in ThreatLocker API',
                    'computer_id': computer_id,
                    'hostname': hostname
                }), 404
            
            # Normalize the device data
            normalized = normalize_threatlocker_device(target_device)
            
            # Get today's date for snapshot
            snapshot_date = date.today()
            
            # Delete existing snapshots for this device on today's date
            delete_query = text("""
                DELETE FROM device_snapshot 
                WHERE snapshot_date = :snapshot_date 
                AND vendor_id = :vendor_id 
                AND device_identity_id IN (
                    SELECT id FROM device_identity 
                    WHERE vendor_id = :vendor_id 
                    AND vendor_device_key = :vendor_device_key
                )
            """)
            
            session.execute(delete_query, {
                'snapshot_date': snapshot_date,
                'vendor_id': vendor_id,
                'vendor_device_key': normalized['vendor_device_key']
            })
            
            # Create or update device identity
            device_identity_id = upsert_device_identity(
                session=session,
                vendor_id=vendor_id,
                vendor_device_key=normalized['vendor_device_key'],
                first_seen_date=snapshot_date
            )
            
            # Insert the updated snapshot
            insert_snapshot(
                session=session,
                snapshot_date=snapshot_date,
                vendor_id=vendor_id,
                device_identity_id=device_identity_id,
                normalized=normalized
            )
            
            session.commit()
            
            return jsonify({
                'success': True,
                'message': 'ThreatLocker device synced successfully',
                'device': {
                    'computer_id': target_device.get('computerId'),
                    'hostname': normalized.get('hostname'),
                    'display_name': normalized.get('display_name'),
                    'organization_name': normalized.get('organization_name')
                },
                'updated_by': updated_by,
                'updated_at': datetime.now().isoformat()
            })
            
    except Exception as e:
        return jsonify({
            'error': f'Failed to sync ThreatLocker device: {str(e)}'
        }), 500

@app.route('/api/variance-report/filtered', methods=['GET'])
def get_filtered_variance_report():
    """
    Get filtered variance report for dashboard integration.
    
    This endpoint provides the same data format as the Variances dashboard
    but uses Database AI's filtered exception data (only unresolved exceptions).
    This ensures both systems use the same authoritative data source.
    """
    latest_date = get_latest_matching_date()
    
    if not latest_date:
        return jsonify({
            "error": "No matching data found between vendors",
            "status": "out_of_sync"
        }), 400
    
    with get_session() as session:
        # Get device counts by vendor
        device_counts_query = text("""
            SELECT v.name as vendor, COUNT(*) as count
            FROM device_snapshot ds
            JOIN vendor v ON ds.vendor_id = v.id
            WHERE ds.snapshot_date = :report_date
            GROUP BY v.name
        """)
        
        device_counts = session.execute(device_counts_query, {'report_date': latest_date}).fetchall()
        device_counts_dict = {row[0]: row[1] for row in device_counts}
        
        # Get only unresolved exceptions (filtered data)
        exceptions_query = text("""
            SELECT id, type, hostname, details
            FROM exceptions
            WHERE date_found = :report_date
            AND resolved = FALSE
            ORDER BY type, hostname
        """)
        
        exceptions = session.execute(exceptions_query, {'report_date': latest_date}).fetchall()
        
        # Group exceptions by type and organization
        by_type = {}
        by_organization = {}
        
        for exc in exceptions:
            exc_type = exc[1]
            hostname = exc[2]
            details = exc[3]
            
            # Initialize type group
            if exc_type not in by_type:
                by_type[exc_type] = []
            
            # Get organization from details
            org_name = details.get('tl_org_name', details.get('ninja_org_name', 'Unknown'))
            if org_name not in by_organization:
                by_organization[org_name] = []
            
            # Format exception data for dashboard
            exception_data = {
                "hostname": hostname,
                "details": details,
                "action": _get_action_for_exception_type(exc_type)
            }
            
            # Add type-specific fields
            if exc_type == 'DISPLAY_NAME_MISMATCH':
                exception_data.update({
                    "ninja_display_name": details.get('ninja_display_name', ''),
                    "threatlocker_computer_name": details.get('tl_display_name', ''),
                    "organization": org_name
                })
            elif exc_type == 'MISSING_NINJA':
                exception_data.update({
                    "threatlocker_hostname": details.get('tl_hostname', ''),
                    "organization": org_name
                })
            
            by_type[exc_type].append(exception_data)
            by_organization[org_name].append(exception_data)
        
        # Calculate totals
        total_variances = len(exceptions)
        
        # Build response in dashboard format
        response = {
            "analysis_date": latest_date.isoformat(),
            "total_devices": {
                "ninja": device_counts_dict.get('Ninja', 0),
                "threatlocker": device_counts_dict.get('ThreatLocker', 0)
            },
            "total_variances": total_variances,
            "data_status": get_data_status(),
            "status": "current"
        }
        
        # Add exception data by type
        if 'DISPLAY_NAME_MISMATCH' in by_type:
            response["display_name_mismatches"] = {
                "total_count": len(by_type['DISPLAY_NAME_MISMATCH']),
                "by_organization": _group_by_organization(by_type['DISPLAY_NAME_MISMATCH'])
            }
        
        if 'MISSING_NINJA' in by_type:
            response["missing_in_ninja"] = {
                "total_count": len(by_type['MISSING_NINJA']),
                "by_organization": _group_by_organization(by_type['MISSING_NINJA'])
            }
        
        if 'DUPLICATE_TL' in by_type:
            response["threatlocker_duplicates"] = {
                "total_count": len(by_type['DUPLICATE_TL']),
                "devices": by_type['DUPLICATE_TL']
            }
        
        if 'SPARE_MISMATCH' in by_type:
            response["devices_that_should_not_have_threatlocker"] = {
                "total_count": len(by_type['SPARE_MISMATCH']),
                "by_organization": _group_by_organization(by_type['SPARE_MISMATCH'])
            }
        
        # Add actionable insights
        response["actionable_insights"] = _generate_actionable_insights(by_type)
        
        # Add collection info with timezone-aware timestamps
        collection_info = _get_collection_timestamps(session, latest_date)
        response["collection_info"] = collection_info
        
        # Add data quality indicators
        response["data_quality"] = {
            "total_exceptions": total_variances,
            "exception_types": list(by_type.keys()),
            "organizations_affected": len(by_organization)
        }
        
        return jsonify(response)

def _get_action_for_exception_type(exc_type: str) -> str:
    """Get recommended action for exception type."""
    actions = {
        'DISPLAY_NAME_MISMATCH': 'Investigate - Reconcile naming differences',
        'MISSING_NINJA': 'Add device to Ninja or remove from ThreatLocker',
        'DUPLICATE_TL': 'Remove duplicate ThreatLocker entries',
        'SPARE_MISMATCH': 'Remove ThreatLocker from spare devices or update billing status',
        'SITE_MISMATCH': 'Reconcile site assignments between vendors'
    }
    return actions.get(exc_type, 'Investigate and resolve')

def _group_by_organization(exceptions: list) -> dict:
    """Group exceptions by organization."""
    by_org = {}
    for exc in exceptions:
        org = exc.get('organization', 'Unknown')
        if org not in by_org:
            by_org[org] = []
        by_org[org].append(exc)
    return by_org

def _generate_actionable_insights(by_type: dict) -> dict:
    """Generate actionable insights based on exception data."""
    insights = {
        "priority_actions": [],
        "summary": {}
    }
    
    for exc_type, exceptions in by_type.items():
        count = len(exceptions)
        if count > 0:
            insights["summary"][exc_type] = count
            
            if exc_type == 'DISPLAY_NAME_MISMATCH' and count > 50:
                insights["priority_actions"].append(f"High priority: {count} display name mismatches need attention")
            elif exc_type == 'MISSING_NINJA' and count > 20:
                insights["priority_actions"].append(f"Critical: {count} devices missing from Ninja")
    
    return insights


def _get_collection_timestamps(session, latest_date: date) -> dict:
    """Get timezone-aware collection timestamps from JobRuns table."""
    from sqlalchemy import text
    
    # Query for the most recent successful collection runs
    collection_query = text("""
        SELECT job_name, started_at, ended_at
        FROM job_runs 
        WHERE status = 'completed'
        AND job_name IN ('ninja-collector', 'threatlocker-collector')
        ORDER BY started_at DESC
    """)
    
    results = session.execute(collection_query).fetchall()
    
    # Initialize timestamps
    ninja_collected = None
    threatlocker_collected = None
    last_collection = None
    
    # Process results - take the most recent for each collector type
    for job_name, started_at, ended_at in results:
        # Use ended_at if available, otherwise started_at
        timestamp = ended_at if ended_at else started_at
        
        # Format timestamp properly (ISO 8601 with Z suffix)
        # Remove any existing timezone info and add Z for UTC
        if timestamp.tzinfo is not None:
            # Convert to UTC and format
            utc_timestamp = timestamp.astimezone().replace(tzinfo=None)
            formatted_timestamp = utc_timestamp.isoformat() + 'Z'
        else:
            # Already UTC, just add Z
            formatted_timestamp = timestamp.isoformat() + 'Z'
        
        # Only set if we haven't found a timestamp for this collector yet
        # (since results are ordered by started_at DESC, first occurrence is most recent)
        if job_name == 'ninja-collector' and ninja_collected is None:
            ninja_collected = formatted_timestamp
        elif job_name == 'threatlocker-collector' and threatlocker_collected is None:
            threatlocker_collected = formatted_timestamp
    
    # Determine the latest collection time
    timestamps = [t for t in [ninja_collected, threatlocker_collected] if t]
    if timestamps:
        # Parse timestamps and find the latest
        parsed_times = []
        for ts in timestamps:
            try:
                # Remove Z suffix and parse as datetime
                clean_ts = ts.replace('Z', '')
                parsed_times.append(datetime.fromisoformat(clean_ts))
            except:
                continue
        
        if parsed_times:
            last_collection = max(parsed_times).isoformat() + 'Z'
    
    # Fallback to date if no specific timestamps found
    if not last_collection:
        last_collection = latest_date.isoformat()
    
    # Determine data freshness
    today = datetime.now().date()
    days_old = (today - latest_date).days
    data_freshness = "current" if days_old <= 1 else "stale"
    
    return {
        "last_collection": last_collection,
        "ninja_collected": ninja_collected,
        "threatlocker_collected": threatlocker_collected,
        "data_freshness": data_freshness
    }


@app.route('/api/exceptions/count', methods=['GET'])
def get_exceptions_count():
    """
    Get count of unresolved exceptions by type for today.
    
    This endpoint helps the dashboard get accurate counts for display.
    """
    exception_type = request.args.get('type', 'DISPLAY_NAME_MISMATCH')
    
    with get_session() as session:
        count_query = text("""
            SELECT COUNT(*) as total_count
            FROM exceptions 
            WHERE type = :exception_type
            AND resolved = FALSE
            AND date_found = CURRENT_DATE
        """)
        
        total_count = session.execute(count_query, {
            'exception_type': exception_type
        }).scalar()
        
        return jsonify({
            'success': True,
            'type': exception_type,
            'unresolved_count': total_count,
            'date': datetime.now().date().isoformat()
        })

@app.route('/api/exceptions/status-summary', methods=['GET'])
def get_exceptions_status_summary():
    """
    Get summary of exception statuses for dashboard display.
    
    Provides counts by status, type, and recent activity.
    """
    with get_session() as session:
        # Get status summary
        status_query = text("""
            SELECT 
                'active' as status,
                type,
                COUNT(*) as count,
                COUNT(CASE WHEN resolved = true THEN 1 END) as resolved_count
            FROM exceptions
            WHERE date_found = CURRENT_DATE
            GROUP BY type
            ORDER BY status, type
        """)
        
        status_results = session.execute(status_query).fetchall()
        
        # Skip recent updates query for now - columns don't exist
        recent_results = []
        
        # Format results
        status_summary = {}
        for row in status_results:
            status = row[0]
            exc_type = row[1]
            count = row[2]
            resolved = row[3]
            
            if status not in status_summary:
                status_summary[status] = {}
            
            status_summary[status][exc_type] = {
                'total': count,
                'resolved': resolved,
                'unresolved': count - resolved
            }
        
        recent_updates = [{
            'hostname': row[0],
            'type': row[1],
            'updated_by': row[2],
            'updated_at': row[3].isoformat() if row[3] else None,
            'update_type': row[4]
        } for row in recent_results]
        
        return jsonify({
            'status_summary': status_summary,
            'recent_manual_updates': recent_updates,
            'generated_at': datetime.now().isoformat()
        })

@app.route('/api/devices/search', methods=['GET'])
def search_devices():
    """
    Search for devices across vendors with hostname truncation handling.
    
    This endpoint addresses the critical issue where Ninja truncates hostnames to 15 characters
    while ThreatLocker stores full hostnames, making cross-vendor lookups difficult.
    
    Query parameters:
    - q: Search term (hostname or partial hostname)
    - vendor: Optional vendor filter ('ninja' or 'threatlocker')
    - limit: Maximum results (default 50)
    """
    search_term = request.args.get('q', '').strip()
    vendor_filter = request.args.get('vendor', '').strip().lower()
    limit = int(request.args.get('limit', 50))
    
    if not search_term:
        return jsonify({'error': 'Search term (q) is required'}), 400
    
    if limit > 200:
        limit = 200  # Cap at 200 for performance
    
    with get_session() as session:
        # Build the search query with multiple matching strategies
        query = text("""
            SELECT 
                v.name as vendor,
                ds.hostname,
                ds.display_name,
                ds.organization_name,
                ds.snapshot_date,
                -- Show canonical key for debugging
                CASE 
                    WHEN v.name = 'ThreatLocker' THEN 
                        LOWER(LEFT(SPLIT_PART(SPLIT_PART(ds.hostname,'|',1),'.',1),15))
                    ELSE 
                        LOWER(LEFT(SPLIT_PART(ds.hostname,'.',1),15))
                END as canonical_key,
                -- Indicate if hostname is truncated
                CASE 
                    WHEN LENGTH(ds.hostname) = 15 AND v.name = 'Ninja' THEN true
                    ELSE false
                END as is_truncated
            FROM device_snapshot ds
            JOIN vendor v ON ds.vendor_id = v.id
            WHERE ds.snapshot_date = (
                SELECT MAX(snapshot_date) FROM device_snapshot
            )
            AND ds.hostname IS NOT NULL
        """)
        
        params = {}
        
        # Build the complete query with all conditions
        query_str = str(query)
        
        # Add vendor filter if specified
        if vendor_filter in ['ninja', 'threatlocker']:
            query_str += " AND v.name = :vendor_filter"
            params['vendor_filter'] = vendor_filter.title()
        
        # Add search conditions - multiple strategies to handle truncation
        query_str += """
            AND (
                -- Exact match
                ds.hostname ILIKE :search_term || '%'
                -- Contains match
                OR ds.hostname ILIKE '%' || :search_term || '%'
                -- Canonical key match (handles truncation)
                OR (
                    CASE 
                        WHEN v.name = 'ThreatLocker' THEN 
                            LOWER(LEFT(SPLIT_PART(SPLIT_PART(ds.hostname,'|',1),'.',1),15))
                        ELSE 
                            LOWER(LEFT(SPLIT_PART(ds.hostname,'.',1),15))
                    END = LOWER(LEFT(SPLIT_PART(:search_term,'.',1),15))
                )
                -- Prefix match for truncated hostnames
                OR LEFT(ds.hostname, 15) = LEFT(:search_term, 15)
            )
            ORDER BY v.name, ds.hostname LIMIT :limit
        """
        
        # Create final query
        query = text(query_str)
        params['search_term'] = search_term
        params['limit'] = limit
        
        results = session.execute(query, params).fetchall()
        
        # Group results by canonical key to show cross-vendor matches
        grouped_results = {}
        for row in results:
            canonical_key = row[5]  # canonical_key column
            if canonical_key not in grouped_results:
                grouped_results[canonical_key] = []
            
            grouped_results[canonical_key].append({
                'vendor': row[0],
                'hostname': row[1],
                'display_name': row[2],
                'organization_name': row[3],
                'snapshot_date': row[4].isoformat(),
                'canonical_key': canonical_key,
                'is_truncated': row[6]
            })
        
        # Calculate summary statistics
        total_devices = len(results)
        vendors_found = set(row[0] for row in results)
        truncated_count = sum(1 for row in results if row[6])
        
        return jsonify({
            'search_term': search_term,
            'total_results': total_devices,
            'vendors_found': list(vendors_found),
            'truncated_hostnames': truncated_count,
            'grouped_by_canonical_key': grouped_results,
            'warning': 'Some hostnames may be truncated due to Ninja 15-character limit' if truncated_count > 0 else None
        })

# Export Endpoints

@app.route('/api/variances/export/csv', methods=['GET'])
def export_variances_csv():
    """
    Export variance data to CSV format.
    
    Query parameters:
    - date: Specific date (YYYY-MM-DD) or 'latest' (default)
    - include_resolved: Include resolved exceptions (default: false)
    - variance_type: Filter by exception type (optional)
    """
    # Get query parameters
    date_param = request.args.get('date', 'latest')
    include_resolved = request.args.get('include_resolved', 'false').lower() == 'true'
    exception_type = request.args.get('variance_type', request.args.get('type', ''))
    
    # Determine report date
    if date_param == 'latest':
        report_date = get_latest_matching_date()
        if not report_date:
            return jsonify({
                "error": "No matching data found between vendors",
                "status": "out_of_sync"
            }), 400
    else:
        try:
            report_date = datetime.strptime(date_param, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({"error": "Invalid date format. Use YYYY-MM-DD"}), 400
    
    with get_session() as session:
        # Check if data exists for this date
        check_query = text("""
            SELECT COUNT(DISTINCT vendor_id) as vendor_count
            FROM device_snapshot
            WHERE snapshot_date = :report_date
        """)
        
        result = session.execute(check_query, {'report_date': report_date}).fetchone()
        
        if result[0] < 2:
            return jsonify({
                "error": f"Insufficient data for {date_param}. Found {result[0]} vendors, need 2.",
                "status": "insufficient_data"
            }), 400
        
        # Build query for exceptions
        query_str = """
            SELECT 
                e.id,
                e.date_found,
                e.type,
                e.hostname,
                e.details,
                e.resolved,
                e.resolved_date,
                e.resolved_by,
                e.manually_updated_at,
                e.manually_updated_by,
                e.variance_status,
                e.update_type,
                e.old_value,
                e.new_value
            FROM exceptions e
            WHERE e.date_found = :report_date
        """
        
        params = {'report_date': report_date}
        
        # Add filters
        if not include_resolved:
            query_str += " AND e.resolved = FALSE"
        
        if exception_type:
            query_str += " AND e.type = :exception_type"
            params['exception_type'] = exception_type
        
        query_str += " ORDER BY e.type, e.hostname"
        
        query = text(query_str)
        
        exceptions = session.execute(query, params).fetchall()
        
        # Generate CSV content
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow([
            'ID', 'Date Found', 'Type', 'Hostname', 'Resolved', 'Resolved Date',
            'Resolved By', 'Manually Updated At', 'Manually Updated By',
            'Variance Status', 'Update Type', 'Organization', 'Details'
        ])
        
        # Write data rows
        for exc in exceptions:
            details = exc[4] if exc[4] else {}
            org_name = details.get('tl_org_name', details.get('ninja_org_name', 'Unknown'))
            
            writer.writerow([
                exc[0],  # ID
                exc[1].isoformat() if exc[1] else '',  # Date Found
                exc[2],  # Type
                exc[3],  # Hostname
                'Yes' if exc[5] else 'No',  # Resolved
                exc[6].isoformat() if exc[6] else '',  # Resolved Date
                exc[7] or '',  # Resolved By
                exc[8].isoformat() if exc[8] else '',  # Manually Updated At
                exc[9] or '',  # Manually Updated By
                exc[10] or 'active',  # Variance Status
                exc[11] or '',  # Update Type
                org_name,  # Organization
                json.dumps(details) if details else ''  # Details
            ])
        
        # Create response
        csv_content = output.getvalue()
        output.close()
        
        # Generate filename
        filename = f"variances-{report_date.isoformat()}.csv"
        if exception_type:
            filename = f"variances-{exception_type.lower()}-{report_date.isoformat()}.csv"
        
        response = Response(
            csv_content,
            mimetype='text/csv',
            headers={
                'Content-Disposition': f'attachment; filename={filename}',
                'Content-Type': 'text/csv; charset=utf-8'
            }
        )
        
        return response

@app.route('/api/variances/available-dates', methods=['GET'])
def get_available_dates():
    """
    Get available analysis dates where both vendors have data.
    
    Returns list of dates with data quality status.
    """
    with get_session() as session:
        # Get dates with both vendors
        query = text("""
            SELECT 
                snapshot_date,
                COUNT(DISTINCT vendor_id) as vendor_count,
                COUNT(DISTINCT CASE WHEN v.name = 'Ninja' THEN ds.vendor_id END) as ninja_count,
                COUNT(DISTINCT CASE WHEN v.name = 'ThreatLocker' THEN ds.vendor_id END) as threatlocker_count
            FROM device_snapshot ds
            JOIN vendor v ON ds.vendor_id = v.id
            WHERE snapshot_date >= CURRENT_DATE - INTERVAL '30 days'
            GROUP BY snapshot_date
            HAVING COUNT(DISTINCT vendor_id) = 2
            ORDER BY snapshot_date DESC
        """)
        
        results = session.execute(query).fetchall()
        
        # Get exception counts for each date
        dates_with_exceptions = []
        for row in results:
            snapshot_date = row[0]
            ninja_count = row[2]
            threatlocker_count = row[3]
            
            # Get exception count for this date
            exception_query = text("""
                SELECT 
                    COUNT(*) as total_exceptions,
                    COUNT(CASE WHEN resolved = FALSE THEN 1 END) as unresolved_exceptions
                FROM exceptions
                WHERE date_found = :snapshot_date
            """)
            
            exception_result = session.execute(exception_query, {'snapshot_date': snapshot_date}).fetchone()
            total_exceptions = exception_result[0] if exception_result else 0
            unresolved_exceptions = exception_result[1] if exception_result else 0
            
            # Determine data quality
            days_old = (date.today() - snapshot_date).days
            if days_old <= 1:
                quality_status = "current"
            elif days_old <= 3:
                quality_status = "recent"
            else:
                quality_status = "stale"
            
            dates_with_exceptions.append({
                'date': snapshot_date.isoformat(),
                'ninja_devices': ninja_count,
                'threatlocker_devices': threatlocker_count,
                'total_exceptions': total_exceptions,
                'unresolved_exceptions': unresolved_exceptions,
                'data_quality': quality_status,
                'days_old': days_old
            })
        
        return jsonify({
            'available_dates': dates_with_exceptions,
            'date_range': {
                'earliest': dates_with_exceptions[-1]['date'] if dates_with_exceptions else None,
                'latest': dates_with_exceptions[0]['date'] if dates_with_exceptions else None
            },
            'total_dates': len(dates_with_exceptions)
        })

@app.route('/api/variances/historical/<date_str>', methods=['GET'])
def get_historical_variance_data(date_str: str):
    """
    Get variance data for a specific historical date.
    
    Same structure as /api/variance-report/latest but for historical date.
    """
    # Get query parameters
    include_resolved = request.args.get('include_resolved', 'false').lower() == 'true'
    
    try:
        report_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({"error": "Invalid date format. Use YYYY-MM-DD"}), 400
    
    with get_session() as session:
        # Check if data exists for this date
        check_query = text("""
            SELECT COUNT(DISTINCT vendor_id) as vendor_count
            FROM device_snapshot
            WHERE snapshot_date = :report_date
        """)
        
        result = session.execute(check_query, {'report_date': report_date}).fetchone()
        
        if result[0] < 2:
            return jsonify({
                "error": f"Insufficient data for {date_str}. Found {result[0]} vendors, need 2.",
                "status": "insufficient_data"
            }), 400
        
        # Get exceptions for the date
        if include_resolved:
            query = text("""
                SELECT id, date_found, type, hostname, details, resolved
                FROM exceptions
                WHERE date_found = :report_date
                ORDER BY type, hostname
            """)
        else:
            query = text("""
                SELECT id, date_found, type, hostname, details, resolved
                FROM exceptions
                WHERE date_found = :report_date
                AND resolved = FALSE
                ORDER BY type, hostname
            """)
        
        exceptions = session.execute(query, {'report_date': report_date}).fetchall()
        
        # Group by type
        by_type = {}
        for exc in exceptions:
            exc_type = exc[2]
            if exc_type not in by_type:
                by_type[exc_type] = []
            by_type[exc_type].append({
                'id': exc[0],
                'hostname': exc[3],
                'details': exc[4],
                'resolved': exc[5]
            })
        
        # Calculate totals
        total_exceptions = len(exceptions)
        unresolved_count = sum(1 for exc in exceptions if not exc[5])
        
        # Get device counts for this date
        device_query = text("""
            SELECT v.name as vendor, COUNT(*) as count
            FROM device_snapshot ds
            JOIN vendor v ON ds.vendor_id = v.id
            WHERE ds.snapshot_date = :report_date
            GROUP BY v.name
        """)
        
        device_counts = {row[0]: row[1] for row in session.execute(device_query, {'report_date': report_date}).fetchall()}
        
        # Get detailed organization breakdown
        org_breakdown = get_organization_breakdown(session, exceptions, report_date)
        
        # Create enhanced format with organization data
        enhanced_format = {
            "missing_in_ninja": org_breakdown.get("missing_in_ninja", {
                "total_count": 0,
                "by_organization": {}
            }),
            "threatlocker_duplicates": org_breakdown.get("threatlocker_duplicates", {
                "total_count": 0,
                "by_organization": {}
            }),
            "ninja_duplicates": org_breakdown.get("ninja_duplicates", {
                "total_count": 0,
                "by_organization": {}
            }),
            "display_name_mismatches": org_breakdown.get("display_name_mismatches", {
                "total_count": 0,
                "by_organization": {}
            })
        }
        
        return jsonify({
            "report_date": report_date.isoformat(),
            "summary": {
                "total_exceptions": total_exceptions,
                "unresolved_count": unresolved_count,
                "resolved_count": total_exceptions - unresolved_count
            },
            "exceptions_by_type": by_type,
            "exception_counts": {exc_type: len(devices) for exc_type, devices in by_type.items()},
            "device_counts": device_counts,
            "data_status": {
                "status": "historical",
                "message": f"Historical data for {date_str}",
                "latest_date": report_date.isoformat()
            },
            # Enhanced format with organization breakdown
            **enhanced_format
        })

@app.route('/api/variances/trends', methods=['GET'])
def get_variance_trends():
    """
    Get trend analysis for variance data over time.
    
    Query parameters:
    - start_date: Start date (YYYY-MM-DD)
    - end_date: End date (YYYY-MM-DD)
    - type: Exception type filter (optional)
    """
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    exception_type = request.args.get('type', '')
    
    if not start_date_str or not end_date_str:
        return jsonify({"error": "start_date and end_date are required"}), 400
    
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({"error": "Invalid date format. Use YYYY-MM-DD"}), 400
    
    if start_date > end_date:
        return jsonify({"error": "start_date must be before end_date"}), 400
    
    with get_session() as session:
        # Build trend query
        query = text("""
            SELECT 
                date_found,
                type,
                COUNT(*) as total_count,
                COUNT(CASE WHEN resolved = FALSE THEN 1 END) as unresolved_count,
                COUNT(CASE WHEN resolved = TRUE THEN 1 END) as resolved_count
            FROM exceptions
            WHERE date_found BETWEEN :start_date AND :end_date
        """)
        
        params = {'start_date': start_date, 'end_date': end_date}
        
        if exception_type:
            query += " AND type = :exception_type"
            params['exception_type'] = exception_type
        
        query += """
            GROUP BY date_found, type
            ORDER BY date_found DESC, type
        """
        
        results = session.execute(query, params).fetchall()
        
        # Group by date and type
        trends_by_date = {}
        trends_by_type = {}
        
        for row in results:
            date_found = row[0]
            exc_type = row[1]
            total_count = row[2]
            unresolved_count = row[3]
            resolved_count = row[4]
            
            # Group by date
            if date_found not in trends_by_date:
                trends_by_date[date_found] = {}
            trends_by_date[date_found][exc_type] = {
                'total': total_count,
                'unresolved': unresolved_count,
                'resolved': resolved_count
            }
            
            # Group by type
            if exc_type not in trends_by_type:
                trends_by_type[exc_type] = []
            trends_by_type[exc_type].append({
                'date': date_found.isoformat(),
                'total': total_count,
                'unresolved': unresolved_count,
                'resolved': resolved_count
            })
        
        # Calculate summary statistics
        total_exceptions = sum(row[2] for row in results)
        total_unresolved = sum(row[3] for row in results)
        total_resolved = sum(row[4] for row in results)
        
        # Get unique dates and types
        unique_dates = sorted(set(row[0] for row in results), reverse=True)
        unique_types = sorted(set(row[1] for row in results))
        
        return jsonify({
            'trends_by_date': {date.isoformat(): data for date, data in trends_by_date.items()},
            'trends_by_type': trends_by_type,
            'summary': {
                'total_exceptions': total_exceptions,
                'total_unresolved': total_unresolved,
                'total_resolved': total_resolved,
                'date_range': {
                    'start': start_date.isoformat(),
                    'end': end_date.isoformat()
                },
                'unique_dates': len(unique_dates),
                'unique_types': unique_types
            },
            'date_range': {
                'start_date': start_date.isoformat(),
                'end_date': end_date.isoformat(),
                'days': (end_date - start_date).days + 1
            }
        })

# Enhanced Export Endpoints

@app.route('/api/variances/export/pdf', methods=['GET'])
def export_variances_pdf():
    """
    Export variance data to PDF format.
    
    Query parameters:
    - date: Specific date (YYYY-MM-DD) or 'latest' (default)
    - include_resolved: Include resolved exceptions (default: false)
    - variance_type: Filter by exception type (optional)
    """
    if not REPORTLAB_AVAILABLE:
        return jsonify({
            "error": "PDF export not available. Install reportlab: pip install reportlab"
        }), 500
    
    # Get query parameters
    date_param = request.args.get('date', 'latest')
    include_resolved = request.args.get('include_resolved', 'false').lower() == 'true'
    exception_type = request.args.get('variance_type', request.args.get('type', ''))
    
    # Determine report date
    if date_param == 'latest':
        report_date = get_latest_matching_date()
        if not report_date:
            return jsonify({
                "error": "No matching data found between vendors",
                "status": "out_of_sync"
            }), 400
    else:
        try:
            report_date = datetime.strptime(date_param, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({"error": "Invalid date format. Use YYYY-MM-DD"}), 400
    
    with get_session() as session:
        # Check if data exists for this date
        check_query = text("""
            SELECT COUNT(DISTINCT vendor_id) as vendor_count
            FROM device_snapshot
            WHERE snapshot_date = :report_date
        """)
        
        result = session.execute(check_query, {'report_date': report_date}).fetchone()
        
        if result[0] < 2:
            return jsonify({
                "error": f"Insufficient data for {date_param}. Found {result[0]} vendors, need 2.",
                "status": "insufficient_data"
            }), 400
        
        # Get exceptions data
        query_str = """
            SELECT 
                e.id,
                e.date_found,
                e.type,
                e.hostname,
                e.details,
                e.resolved,
                e.resolved_date,
                e.resolved_by,
                e.manually_updated_at,
                e.manually_updated_by,
                e.variance_status
            FROM exceptions e
            WHERE e.date_found = :report_date
        """
        
        params = {'report_date': report_date}
        
        # Add filters
        if not include_resolved:
            query_str += " AND e.resolved = FALSE"
        
        if exception_type:
            query_str += " AND e.type = :exception_type"
            params['exception_type'] = exception_type
        
        query_str += " ORDER BY e.type, e.hostname"
        
        query = text(query_str)
        
        exceptions = session.execute(query, params).fetchall()
        
        # Generate PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center
        )
        story.append(Paragraph(f"Variance Report - {report_date.isoformat()}", title_style))
        story.append(Spacer(1, 12))
        
        # Summary
        total_exceptions = len(exceptions)
        unresolved_count = sum(1 for exc in exceptions if not exc[5])
        resolved_count = total_exceptions - unresolved_count
        
        summary_data = [
            ['Total Exceptions', str(total_exceptions)],
            ['Unresolved', str(unresolved_count)],
            ['Resolved', str(resolved_count)],
            ['Report Date', report_date.isoformat()],
            ['Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
        ]
        
        summary_table = Table(summary_data, colWidths=[2*inch, 1.5*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(Paragraph("Summary", styles['Heading2']))
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Exceptions table
        if exceptions:
            story.append(Paragraph("Exception Details", styles['Heading2']))
            
            # Prepare table data
            table_data = [['ID', 'Type', 'Hostname', 'Status', 'Organization']]
            
            for exc in exceptions:
                details = exc[4] if exc[4] else {}
                org_name = details.get('tl_org_name', details.get('ninja_org_name', 'Unknown'))
                status = 'Resolved' if exc[5] else 'Active'
                
                table_data.append([
                    str(exc[0]),
                    exc[2],
                    exc[3],
                    status,
                    org_name
                ])
            
            # Create table
            exceptions_table = Table(table_data, colWidths=[0.5*inch, 1.2*inch, 1.5*inch, 0.8*inch, 1.5*inch])
            exceptions_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.white])
            ]))
            
            story.append(exceptions_table)
        else:
            story.append(Paragraph("No exceptions found for the specified criteria.", styles['Normal']))
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        
        # Generate filename
        filename = f"variances-{report_date.isoformat()}.pdf"
        if exception_type:
            filename = f"variances-{exception_type.lower()}-{report_date.isoformat()}.pdf"
        
        return Response(
            buffer.getvalue(),
            mimetype='application/pdf',
            headers={
                'Content-Disposition': f'attachment; filename={filename}',
                'Content-Type': 'application/pdf'
            }
        )

@app.route('/api/variances/export/excel', methods=['GET'])
def export_variances_excel():
    """
    Export variance data to Excel format with multiple sheets.
    
    Query parameters:
    - date: Specific date (YYYY-MM-DD) or 'latest' (default)
    - include_resolved: Include resolved exceptions (default: false)
    - variance_type: Filter by exception type (optional)
    """
    if not OPENPYXL_AVAILABLE:
        return jsonify({
            "error": "Excel export not available. Install openpyxl: pip install openpyxl"
        }), 500
    
    # Get query parameters
    date_param = request.args.get('date', 'latest')
    include_resolved = request.args.get('include_resolved', 'false').lower() == 'true'
    exception_type = request.args.get('variance_type', request.args.get('type', ''))
    
    # Determine report date
    if date_param == 'latest':
        report_date = get_latest_matching_date()
        if not report_date:
            return jsonify({
                "error": "No matching data found between vendors",
                "status": "out_of_sync"
            }), 400
    else:
        try:
            report_date = datetime.strptime(date_param, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({"error": "Invalid date format. Use YYYY-MM-DD"}), 400
    
    with get_session() as session:
        # Check if data exists for this date
        check_query = text("""
            SELECT COUNT(DISTINCT vendor_id) as vendor_count
            FROM device_snapshot
            WHERE snapshot_date = :report_date
        """)
        
        result = session.execute(check_query, {'report_date': report_date}).fetchone()
        
        if result[0] < 2:
            return jsonify({
                "error": f"Insufficient data for {date_param}. Found {result[0]} vendors, need 2.",
                "status": "insufficient_data"
            }), 400
        
        # Get exceptions data
        query_str = """
            SELECT 
                e.id,
                e.date_found,
                e.type,
                e.hostname,
                e.details,
                e.resolved,
                e.resolved_date,
                e.resolved_by,
                e.manually_updated_at,
                e.manually_updated_by,
                e.variance_status,
                e.update_type
            FROM exceptions e
            WHERE e.date_found = :report_date
        """
        
        params = {'report_date': report_date}
        
        # Add filters
        if not include_resolved:
            query_str += " AND e.resolved = FALSE"
        
        if exception_type:
            query_str += " AND e.type = :exception_type"
            params['exception_type'] = exception_type
        
        query_str += " ORDER BY e.type, e.hostname"
        
        query = text(query_str)
        
        exceptions = session.execute(query, params).fetchall()
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create Summary sheet
        summary_ws = wb.create_sheet("Summary")
        summary_ws.title = "Summary"
        
        # Summary data
        total_exceptions = len(exceptions)
        unresolved_count = sum(1 for exc in exceptions if not exc[5])
        resolved_count = total_exceptions - unresolved_count
        
        summary_data = [
            ['Variance Report Summary'],
            [''],
            ['Report Date', report_date.isoformat()],
            ['Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            [''],
            ['Total Exceptions', total_exceptions],
            ['Unresolved', unresolved_count],
            ['Resolved', resolved_count],
            [''],
            ['Exception Types:']
        ]
        
        # Add exception type breakdown
        type_counts = {}
        for exc in exceptions:
            exc_type = exc[2]
            type_counts[exc_type] = type_counts.get(exc_type, 0) + 1
        
        for exc_type, count in sorted(type_counts.items()):
            summary_data.append([exc_type, count])
        
        # Write summary data
        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, cell_value in enumerate(row_data, 1):
                cell = summary_ws.cell(row=row_idx, column=col_idx, value=cell_value)
                if row_idx == 1:  # Title
                    cell.font = Font(bold=True, size=14)
                elif row_idx in [3, 4, 6, 7, 8]:  # Data rows
                    cell.font = Font(bold=True)
        
        # Create detailed exceptions sheet
        details_ws = wb.create_sheet("Exceptions")
        details_ws.title = "Exceptions"
        
        # Headers
        headers = [
            'ID', 'Date Found', 'Type', 'Hostname', 'Resolved',
            'Resolved Date', 'Resolved By', 'Manually Updated At',
            'Manually Updated By', 'Variance Status', 'Update Type', 'Organization'
        ]
        
        for col_idx, header in enumerate(headers, 1):
            cell = details_ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data rows
        for row_idx, exc in enumerate(exceptions, 2):
            details = exc[4] if exc[4] else {}
            org_name = details.get('tl_org_name', details.get('ninja_org_name', 'Unknown'))
            
            row_data = [
                exc[0],  # ID
                exc[1].isoformat() if exc[1] else '',  # Date Found
                exc[2],  # Type
                exc[3],  # Hostname
                'Yes' if exc[5] else 'No',  # Resolved
                exc[6].isoformat() if exc[6] else '',  # Resolved Date
                exc[7] or '',  # Resolved By
                exc[8].isoformat() if exc[8] else '',  # Manually Updated At
                exc[9] or '',  # Manually Updated By
                exc[10] or 'active',  # Variance Status
                exc[11] or '',  # Update Type
                org_name  # Organization
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                details_ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for ws in [summary_ws, details_ws]:
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Generate filename
        filename = f"variances-{report_date.isoformat()}.xlsx"
        if exception_type:
            filename = f"variances-{exception_type.lower()}-{report_date.isoformat()}.xlsx"
        
        return Response(
            buffer.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename={filename}',
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
        )


# Windows 11 24H2 Assessment Endpoints
@app.route('/api/windows-11-24h2/status', methods=['GET'])
def get_windows_11_24h2_status():
    """Get Windows 11 24H2 compatibility status summary"""
    try:
        with get_session() as session:
            # Query database for Windows 11 24H2 assessment results with enhanced status breakdown
            # CRITICAL: Do NOT filter by windows_11_24h2_capable IS NOT NULL - we need to count ALL Windows devices
            # including those not yet assessed (which will have NULL values)
            query = text("""
            SELECT 
                COUNT(*) as total_windows_devices,
                COUNT(CASE WHEN windows_11_24h2_capable = true THEN 1 END) as total_compatible_devices,
                COUNT(CASE WHEN windows_11_24h2_capable = false THEN 1 END) as incompatible_devices,
                COUNT(CASE WHEN windows_11_24h2_capable IS NULL THEN 1 END) as not_assessed_devices,
                COUNT(CASE WHEN windows_11_24h2_capable = true 
                          AND windows_11_24h2_deficiencies::jsonb->>'passed_requirements' LIKE '%Windows 11 24H2 Already Installed%' 
                          THEN 1 END) as already_compatible_devices,
                COUNT(CASE WHEN windows_11_24h2_capable = true 
                          AND (windows_11_24h2_deficiencies::jsonb->>'passed_requirements' NOT LIKE '%Windows 11 24H2 Already Installed%' 
                               OR windows_11_24h2_deficiencies::jsonb->>'passed_requirements' IS NULL)
                          THEN 1 END) as compatible_for_upgrade_devices
            FROM device_snapshot ds
            JOIN vendor v ON ds.vendor_id = v.id
            JOIN device_type dt ON ds.device_type_id = dt.id
            WHERE v.name = 'Ninja' 
            AND ds.snapshot_date = (
                SELECT MAX(snapshot_date)
                FROM device_snapshot ds2
                JOIN vendor v2 ON ds2.vendor_id = v2.id
                WHERE v2.name = 'Ninja'
            )
            AND ds.os_name ILIKE '%windows%'
            AND ds.os_name NOT ILIKE '%server%'
            AND dt.code IN ('Desktop', 'Laptop', 'workstation')
            """)
            
            result = session.execute(query).fetchone()
            
            # Get last assessment date
            last_assessment_query = text("""
            SELECT MAX(assessment_date) as last_assessment
            FROM (
                SELECT windows_11_24h2_deficiencies::jsonb->>'assessment_date' as assessment_date
                FROM device_snapshot 
                WHERE windows_11_24h2_deficiencies IS NOT NULL 
                AND windows_11_24h2_deficiencies != '{}'
            ) as assessments
            """)
            
            last_assessment_result = session.execute(last_assessment_query).fetchone()
            last_assessment = last_assessment_result.last_assessment if last_assessment_result else None
            
            return jsonify({
                "total_windows_devices": result.total_windows_devices,
                "total_compatible_devices": result.total_compatible_devices,
                "already_compatible_devices": result.already_compatible_devices,
                "compatible_for_upgrade_devices": result.compatible_for_upgrade_devices,
                "incompatible_devices": result.incompatible_devices,
                "not_assessed_devices": result.not_assessed_devices,
                "compatibility_rate": round((result.total_compatible_devices / result.total_windows_devices * 100), 1) if result.total_windows_devices > 0 else 0,
                "last_assessment": last_assessment
            })
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/windows-11-24h2/incompatible', methods=['GET'])
def get_incompatible_devices():
    """Get list of devices that are incompatible with Windows 11 24H2"""
    try:
        with get_session() as session:
            query = text("""
            SELECT 
                ds.hostname,
                ds.display_name,
                ds.organization_name,
                ds.location_name,
                ds.device_type_name,
                ds.billable_status_name,
                ds.os_name,
                ds.os_build,
                ds.os_release_id,
                ds.memory_gib,
                ds.cpu_model,
                ds.last_online,
                ds.created_at,
                ds.windows_11_24h2_deficiencies,
                ds.system_manufacturer,
                ds.system_model
            FROM device_snapshot ds
            JOIN vendor v ON ds.vendor_id = v.id
            WHERE v.name = 'Ninja'
            AND ds.snapshot_date = (
                SELECT MAX(snapshot_date)
                FROM device_snapshot ds2
                JOIN vendor v2 ON ds2.vendor_id = v2.id
                WHERE v2.name = 'Ninja'
            )
            AND ds.os_name ILIKE '%windows%'
            AND ds.os_name NOT ILIKE '%server%'
            AND ds.windows_11_24h2_capable IS NOT NULL
            AND ds.windows_11_24h2_capable = false
            AND ds.device_type_id IN (SELECT id FROM device_type WHERE code IN ('Desktop', 'Laptop', 'workstation'))
            ORDER BY ds.organization_name, ds.hostname
            """)
            
            results = session.execute(query).fetchall()
            
            devices = []
            for row in results:
                deficiencies = {}
                if row.windows_11_24h2_deficiencies:
                    try:
                        import json
                        deficiencies = json.loads(row.windows_11_24h2_deficiencies)
                    except (json.JSONDecodeError, TypeError):
                        deficiencies = {}
                
                devices.append({
                    # Modal column mapping - using database fields
                    "organization": row.organization_name if row.organization_name and row.organization_name.strip() else "Unknown",
                    "location": row.location_name or "Main Office",
                    "system_name": row.hostname,  # System hostname/identifier
                    "display_name": row.display_name or row.hostname,  # User-friendly device name
                    "device_type": row.device_type_name or "Desktop",  # Physical device type
                    "billable_status": row.billable_status_name or "Active",  # Billing status
                    "status": "Incompatible",  # Clear status
                    
                    # Additional fields for reference
                    "hostname": row.hostname,
                    "os_name": row.os_name,
                    "os_version": row.os_release_id,  # Use os_release_id as version
                    "os_build": row.os_build,
                    "deficiencies": deficiencies.get('deficiencies', []),
                    "assessment_date": _format_date_string(deficiencies.get('assessment_date', 'Unknown')),
                    
                    # New fields requested by Dashboard AI
                    "last_update": row.created_at.replace(microsecond=0).replace(tzinfo=None).isoformat() + 'Z' if row.created_at else None,
                    "last_contact": row.last_online.replace(microsecond=0).replace(tzinfo=None).isoformat() + 'Z' if row.last_online else None,
                    "cpu_model": row.cpu_model or "Unknown",
                    "memory_gib": float(row.memory_gib) if row.memory_gib else None,
                    "system_manufacturer": row.system_manufacturer or "Not Available",
                    "system_model": row.system_model or "Not Available"
                })
            
            return jsonify({
                "incompatible_devices": devices,
                "total_count": len(devices),
                "data_source": "Database (NinjaRMM fields populated)",
                "last_updated": datetime.utcnow().isoformat() + 'Z'
            })
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/windows-11-24h2/compatible', methods=['GET'])
def get_compatible_devices():
    """Get list of devices that are compatible with Windows 11 24H2"""
    try:
        
        with get_session() as session:
            # Get compatible devices from database
            query = text("""
            SELECT 
                ds.hostname,
                ds.display_name,
                ds.organization_name,
                ds.location_name,
                ds.device_type_name,
                ds.billable_status_name,
                ds.os_name,
                ds.os_build,
                ds.os_release_id,
                ds.memory_gib,
                ds.cpu_model,
                ds.last_online,
                ds.created_at,
                ds.windows_11_24h2_deficiencies,
                ds.system_manufacturer,
                ds.system_model
            FROM device_snapshot ds
            JOIN vendor v ON ds.vendor_id = v.id
            WHERE v.name = 'Ninja'
            AND ds.snapshot_date = (
                SELECT MAX(snapshot_date)
                FROM device_snapshot ds2
                JOIN vendor v2 ON ds2.vendor_id = v2.id
                WHERE v2.name = 'Ninja'
            )
            AND ds.os_name ILIKE '%windows%'
            AND ds.os_name NOT ILIKE '%server%'
            AND ds.windows_11_24h2_capable IS NOT NULL
            AND ds.windows_11_24h2_capable = true
            AND (ds.windows_11_24h2_deficiencies::jsonb->>'passed_requirements' NOT LIKE '%Windows 11 24H2 Already Installed%' 
                 OR ds.windows_11_24h2_deficiencies::jsonb->>'passed_requirements' IS NULL)
            AND ds.device_type_id IN (SELECT id FROM device_type WHERE code IN ('Desktop', 'Laptop', 'workstation'))
            ORDER BY ds.organization_name, ds.hostname
            """)
            
            results = session.execute(query).fetchall()
            
            devices = []
            for row in results:
                assessment_data = {}
                if row.windows_11_24h2_deficiencies:
                    try:
                        import json
                        assessment_data = json.loads(row.windows_11_24h2_deficiencies)
                    except (json.JSONDecodeError, TypeError):
                        assessment_data = {}
                
                devices.append({
                    # Modal column mapping - using database fields
                    "organization": row.organization_name if row.organization_name and row.organization_name.strip() else "Unknown",
                    "location": row.location_name or "Main Office",
                    "system_name": row.hostname,  # System hostname/identifier
                    "display_name": row.display_name or row.hostname,  # User-friendly device name
                    "device_type": row.device_type_name or "Desktop",  # Physical device type
                    "billable_status": row.billable_status_name or "Active",  # Billing status
                    "status": "Compatible for Upgrade",  # Clear status
                    
                    # Additional fields for reference
                    "hostname": row.hostname,
                    "os_name": row.os_name,
                    "os_version": row.os_release_id,  # Use os_release_id as version
                    "os_build": row.os_build,
                    "passed_requirements": assessment_data.get('passed_requirements', []),
                    "assessment_date": _format_date_string(assessment_data.get('assessment_date', 'Unknown')),
                    
                    # New fields requested by Dashboard AI
                    "last_update": row.created_at.replace(microsecond=0).replace(tzinfo=None).isoformat() + 'Z' if row.created_at else None,
                    "last_contact": row.last_online.replace(microsecond=0).replace(tzinfo=None).isoformat() + 'Z' if row.last_online else None,
                    "cpu_model": row.cpu_model or "Unknown",
                    "memory_gib": float(row.memory_gib) if row.memory_gib else None,
                    "system_manufacturer": row.system_manufacturer or "Not Available",
                    "system_model": row.system_model or "Not Available"
                })
            
            return jsonify({
                "compatible_devices": devices,
                "total_count": len(devices),
                "data_source": "Database (NinjaRMM fields populated)",
                "last_updated": datetime.utcnow().isoformat() + 'Z'
            })
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/windows-11-24h2/run', methods=['POST'])
def run_windows_11_24h2_assessment():
    """Manually trigger Windows 11 24H2 assessment"""
    try:
        import subprocess
        import os
        
        # Run the assessment script
        script_path = '/opt/es-inventory-hub/collectors/assessments/windows_11_24h2_assessment.py'
        
        if not os.path.exists(script_path):
            return jsonify({"error": "Assessment script not found"}), 404
        
        # Execute the assessment script
        result = subprocess.run([
            '/opt/es-inventory-hub/.venv/bin/python3',
            script_path
        ], capture_output=True, text=True, cwd='/opt/es-inventory-hub')
        
        if result.returncode == 0:
            return jsonify({
                "status": "success",
                "message": "Windows 11 24H2 assessment completed successfully",
                "output": result.stdout,
                "timestamp": datetime.utcnow().isoformat() + 'Z'
            })
        else:
            return jsonify({
                "status": "error",
                "message": "Assessment failed",
                "error": result.stderr,
                "output": result.stdout
            }), 500
            
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# Documentation endpoints
@app.route('/api/docs', methods=['GET'])
def get_documentation():
    """Serve the main integration guide"""
    try:
        from flask import send_from_directory
        docs_dir = '/opt/es-inventory-hub/docs'
        return send_from_directory(docs_dir, 'GUIDE_DATABASE_INTEGRATION.md')
    except Exception as e:
        return jsonify({"error": f"Documentation not available: {str(e)}"}), 404

@app.route('/api/docs/<filename>', methods=['GET'])
def get_doc_file(filename):
    """Serve specific documentation files"""
    try:
        from flask import send_from_directory
        docs_dir = '/opt/es-inventory-hub/docs'
        return send_from_directory(docs_dir, filename)
    except Exception as e:
        return jsonify({"error": f"Documentation file '{filename}' not available: {str(e)}"}), 404


if __name__ == '__main__':
    print("Starting ES Inventory Hub API Server...")
    print("Available endpoints:")
    print("  GET  /api/health - Health check")
    print("  GET  /api/status - System status")
    print("  GET  /api/variance-report/latest - Latest variance report")
    print("  GET  /api/variance-report/{date} - Variance report for specific date")
    print("  GET  /api/variance-report/filtered - Filtered variance report for dashboard")
    print("  POST /api/collectors/run - Trigger collector runs")
    print("  GET  /api/collectors/status - Collector service status")
    print("  GET  /api/collectors/history - Collection history (last 10 runs)")
    print("  GET  /api/collectors/progress - Real-time collection progress")
    print("  GET  /api/exceptions - Get exceptions with filtering")
    print("  POST /api/exceptions/{id}/resolve - Resolve an exception")
    print("  POST /api/exceptions/{id}/mark-manually-fixed - Mark as manually fixed (NEW)")
    print("  POST /api/exceptions/mark-fixed-by-hostname - Mark exceptions fixed by hostname (NEW)")
    print("  POST /api/exceptions/bulk-update - Bulk exception operations (NEW)")
    print("  GET  /api/exceptions/status-summary - Exception status summary (NEW)")
    print("  POST /api/collectors/threatlocker/run - Run ThreatLocker collector (NEW)")
    print("  POST /api/collectors/cross-vendor/run - Run cross-vendor checks (NEW)")
    print("  POST /api/collectors/sequence/run - Run complete collector sequence (NEW)")
    print("  POST /api/threatlocker/update-name - Update ThreatLocker computer name (NEW)")
    print("  POST /api/threatlocker/sync-device - Sync ThreatLocker device to database (NEW)")
    print("  GET  /api/devices/search?q={hostname} - Search devices (handles hostname truncation)")
    print()
    print("NEW EXPORT ENDPOINTS:")
    print("  GET  /api/variances/export/csv - Export variance data to CSV")
    print("  GET  /api/variances/export/pdf - Export variance data to PDF")
    print("  GET  /api/variances/export/excel - Export variance data to Excel")
    print("  GET  /api/variances/available-dates - Get available analysis dates")
    print("  GET  /api/variances/historical/{date} - Get historical variance data")
    print("  GET  /api/variances/trends - Get variance trends over time")
    print()
    print("WINDOWS 11 24H2 ASSESSMENT ENDPOINTS:")
    print("  GET  /api/windows-11-24h2/status - Windows 11 24H2 compatibility status summary")
    print("  GET  /api/windows-11-24h2/incompatible - List of incompatible devices")
    print("  GET  /api/windows-11-24h2/compatible - List of compatible devices")
    print("  POST /api/windows-11-24h2/run - Manually trigger Windows 11 24H2 assessment")
    print()
    print("DOCUMENTATION ENDPOINTS:")
    print("  GET  /api/docs - Main integration documentation")
    print("  GET  /api/docs/<filename> - Specific documentation files")
    print()
    print("Server will run on:")
    print("  HTTP:  http://localhost:5400 (development only)")
    print("  HTTPS: https://db-api.enersystems.com:5400 (production)")
    
    # Check for SSL certificates
    ssl_cert = '/opt/es-inventory-hub/ssl/api.crt'
    ssl_key = '/opt/es-inventory-hub/ssl/api.key'
    
    if os.path.exists(ssl_cert) and os.path.exists(ssl_key):
        print(f"SSL certificates found. Starting HTTPS server...")
        app.run(host='0.0.0.0', port=5400, debug=True, ssl_context=(ssl_cert, ssl_key))
    else:
        print(f"SSL certificates not found. Starting HTTP server...")
        print(f"To enable HTTPS, place SSL certificates at:")
        print(f"  Certificate: {ssl_cert}")
        print(f"  Private Key: {ssl_key}")
        app.run(host='0.0.0.0', port=5400, debug=True)
